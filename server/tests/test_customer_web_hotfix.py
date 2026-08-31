import sqlite3
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.services.batch_exports import batch_picking_filename, batch_picking_workbook, batch_picking_workbook_multi
from app.services.customer_unit_codes import (
    CUSTOMER_UNIT_CODES,
    apply_customer_unit_codes,
    preview_customer_unit_codes,
)
from test_workflows import login, make_client


def _unit_connection(rows):
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    conn.execute(
        "CREATE TABLE units(id TEXT PRIMARY KEY, unit_code TEXT UNIQUE, unit_name TEXT NOT NULL, updated_at TEXT)"
    )
    conn.executemany("INSERT INTO units(id, unit_code, unit_name) VALUES (?, ?, ?)", rows)
    conn.commit()
    return conn


def test_customer_unit_code_preview_and_transactional_backfill_preserve_leading_zeroes():
    rows = [(f"unit-{index}", f"OLD-{index}", f"  {name}  ") for index, (_, name) in enumerate(CUSTOMER_UNIT_CODES, 1)]
    conn = _unit_connection(rows)
    preview = preview_customer_unit_codes(conn)
    assert [item.status for item in preview] == ["MATCHED"] * 21
    assert [item.unit_code for item in preview] == [f"{index:03d}" for index in range(1, 22)]

    conn.execute("BEGIN")
    assert apply_customer_unit_codes(conn, preview) == 21
    conn.commit()
    actual = [row[0] for row in conn.execute("SELECT unit_code FROM units ORDER BY unit_code")]
    assert actual == [f"{index:03d}" for index in range(1, 22)]
    assert "001" in actual and "1" not in actual


@pytest.mark.parametrize(
    ("rows", "expected_status"),
    [
        ([("one", "OLD", CUSTOMER_UNIT_CODES[0][1])], "MISSING"),
        (
            [
                ("one", "OLD-1", CUSTOMER_UNIT_CODES[0][1]),
                ("two", "OLD-2", CUSTOMER_UNIT_CODES[0][1]),
            ],
            "AMBIGUOUS",
        ),
        (
            [
                ("one", "OLD-1", CUSTOMER_UNIT_CODES[0][1]),
                ("owner", "001", "其他单位"),
            ],
            "CODE_CONFLICT",
        ),
    ],
)
def test_customer_unit_code_backfill_rejects_incomplete_ambiguous_or_conflicting_preview(rows, expected_status):
    conn = _unit_connection(rows)
    preview = preview_customer_unit_codes(conn)
    assert expected_status in {item.status for item in preview}
    with pytest.raises(ValueError, match="禁止回填"):
        apply_customer_unit_codes(conn, preview)


def test_unit_api_returns_and_sorts_three_digit_codes_as_strings(tmp_path):
    client = make_client(tmp_path)
    admin = login(client, "root_admin", "StrongPassword123")
    for code, name in (("021", "行宫东派出所"), ("001", "三河市公安局"), ("010", "黄土庄派出所")):
        response = client.post(
            "/api/v1/admin/units",
            headers=admin,
            json={"unit_code": code, "unit_name": name, "default_delivery_point": name},
        )
        assert response.status_code == 200, response.text
        assert response.json()["unit_code"] == code

    listed = client.get("/api/v1/admin/units", headers=admin)
    assert listed.status_code == 200, listed.text
    assert [item["unit_code"] for item in listed.json()] == ["001", "010", "021"]
    duplicate = client.post(
        "/api/v1/admin/units",
        headers=admin,
        json={"unit_code": "001", "unit_name": "重复单位", "default_delivery_point": "测试"},
    )
    assert duplicate.status_code == 409


def _single_unit_aggregation(code: str, name: str):
    item = {
        "product_id": "historical-product",
        "product_name": "历史圆白菜",
        "category": "蔬菜",
        "spec": "历史规格",
        "unit": "斤",
        "requested_quantity": "12.5",
        "actual_quantity": "12.5",
        "quantity": "12.5",
        "subtotal_cents": 0,
    }
    return {
        "batch": {
            "batch_no": "PS20260831-0016",
            "created_at": "2026-08-30 16:30:00",
            "business_date": "2026-08-31",
        },
        "by_unit": [
            {
                "unit_id": f"unit-{code}",
                "unit_code": code,
                "unit_name": name,
                "delivery_point": "历史配送点",
                "order_count": 1,
                "total_cents": 0,
                "items": [item],
            }
        ],
        "document_lines": [item],
    }


@pytest.mark.parametrize(
    ("code", "name"),
    [("001", "三河市公安局"), ("004", "三河特巡警大队"), ("021", "行宫东派出所")],
)
def test_unit_picking_sheets_use_date_code_snapshot_and_keep_internal_batch_number(code, name):
    aggregation = _single_unit_aggregation(code, name)
    workbook = load_workbook(BytesIO(batch_picking_workbook(aggregation)), data_only=True)
    sheet = workbook[f"{code}-蔬菜"]
    total = workbook["总计"]

    assert workbook.sheetnames == [f"{code}-蔬菜", "总计"]
    assert sheet["A1"].value == f"蔬菜备货单（20260831/{code}）"
    assert total["A1"].value == f"三公鲜配备货单（20260831/{code}）"
    assert total["A2"].value is None
    assert total["D2"].value == "系统备货单号：PS20260831-0016"
    assert sheet["A2"].value is None
    assert sheet["D2"].value == "系统备货单号：PS20260831-0016"
    assert [sheet.cell(4, column).value for column in range(1, 6)] == [1, "历史圆白菜", "历史规格", "斤", "12.5"]
    assert str(sheet.print_area).endswith("!$A$1:$E$6")
    assert sheet.page_setup.fitToWidth == 1
    assert batch_picking_filename(aggregation) == f"蔬菜备货单_{code}_{name}_20260831.xlsx"

    bulk_workbook = load_workbook(BytesIO(batch_picking_workbook_multi([aggregation])), data_only=True)
    assert bulk_workbook.sheetnames == [f"0016-{code}-蔬菜", "总计-PS20260831-0016"]
    assert bulk_workbook["总计-PS20260831-0016"]["A1"].value == f"三公鲜配备货单（20260831/{code}）"
