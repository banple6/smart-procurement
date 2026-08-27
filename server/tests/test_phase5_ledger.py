from io import BytesIO
from time import perf_counter
from urllib.parse import unquote
from uuid import uuid4

from openpyxl import load_workbook

from test_new_business_features import create_product, create_unit_order, login, make_client


def seed_ledger_data(tmp_path):
    client = make_client(tmp_path)
    admin_headers = login(client, "root_admin", "StrongPassword123")
    potato = create_product(client, admin_headers, "LEDGER-POTATO", "土豆", "斤", 321, "蔬菜")
    boxed_potato = create_product(client, admin_headers, "LEDGER-POTATO-BOX", "土豆", "箱", 500, "蔬菜")
    first = create_unit_order(client, admin_headers, potato["id"], "LEDGER-A", "3")
    first_unit_headers = login(client, "unit_LEDGER-A", "UnitLEDGER-APassword123")
    second_response = client.post(
        "/api/v1/orders",
        headers=first_unit_headers,
        json={"items": [{"product_id": boxed_potato["id"], "quantity": "2"}]},
    )
    assert second_response.status_code == 200, second_response.text
    second = second_response.json()
    third = create_unit_order(client, admin_headers, potato["id"], "LEDGER-B", "1")

    from app.database import connect

    with connect() as conn:
        conn.execute(
            """
            UPDATE orders
            SET status = 'completed', archived_at = '2026-08-05 00:00:00',
                created_at = '2026-08-01 15:59:59', completed_at = '2026-08-01 16:30:00'
            WHERE id = ?
            """,
            (first["id"],),
        )
        conn.execute(
            """
            UPDATE orders
            SET status = 'shipped', archived_at = '2026-08-05 00:00:00',
                created_at = '2026-08-01 16:00:00', shipped_at = '2026-08-01 17:00:00'
            WHERE id = ?
            """,
            (second["id"],),
        )
        conn.execute(
            """
            UPDATE orders
            SET status = 'preparing', created_at = '2026-08-02 03:30:00', preparing_at = '2026-08-02 04:00:00'
            WHERE id = ?
            """,
            (third["id"],),
        )
        conn.commit()

    current_product = client.get(f"/api/v1/products/{potato['id']}", headers=admin_headers).json()
    changed = client.patch(
        f"/api/v1/admin/products/{potato['id']}/price",
        headers=admin_headers,
        json={"price_cents": 1088, "expected_version": current_product["version"]},
    )
    assert changed.status_code == 200, changed.text
    return client, admin_headers, first, second, third


def worksheet_rows(sheet):
    return list(sheet.iter_rows(values_only=True))


def test_phase5_ledger_filters_snapshots_archives_and_exports(tmp_path):
    client, headers, first, second, third = seed_ledger_data(tmp_path)

    legacy = client.get("/api/v1/admin/ledger", headers=headers)
    assert legacy.status_code == 200, legacy.text
    rows = legacy.json()
    assert {row["order_no"] for row in rows} == {first["order_no"], second["order_no"], third["order_no"]}
    first_row = next(row for row in rows if row["order_no"] == first["order_no"])
    assert first_row["status"] == "completed"
    assert first_row["price_cents_snapshot"] == 321
    assert first_row["subtotal_cents"] == 963
    assert first_row["created_at"] == "2026-08-01 23:59:59"

    def page(query=""):
        suffix = f"&{query}" if query else ""
        response = client.get(f"/api/v1/admin/ledger?page=1&page_size=20{suffix}", headers=headers)
        assert response.status_code == 200, response.text
        return response.json()

    first_unit_id = first["unit_id"]
    assert page("start_date=2026-08-01")["total"] == 3
    assert page("end_date=2026-08-01")["total"] == 1
    assert page("start_date=2026-08-01&end_date=2026-08-01")["total"] == 1
    assert page("start_date=2026-08-02&end_date=2026-08-02")["total"] == 2
    assert page(f"unit_id={first_unit_id}")["total"] == 2
    assert page("status=shipped")["items"][0]["order_no"] == second["order_no"]
    assert page(f"order_no=%20{first['order_no']}%20")["items"][0]["order_no"] == first["order_no"]
    assert page(f"start_date=2026-08-02&unit_id={first_unit_id}")["items"][0]["order_no"] == second["order_no"]
    assert page(f"start_date=2026-08-02&unit_id={first_unit_id}&status=shipped")["total"] == 1
    assert page("unit_id=does-not-exist")["total"] == 0
    invalid = client.get("/api/v1/admin/ledger?start_date=2026-08-03&end_date=2026-08-02", headers=headers)
    assert invalid.status_code == 422
    assert invalid.json()["detail"] == "开始日期不能晚于结束日期"

    paged = client.get("/api/v1/admin/ledger?page=1&page_size=1", headers=headers)
    assert paged.status_code == 200
    assert paged.json()["total"] == 3
    assert len(paged.json()["items"]) == 1

    current = client.get(
        f"/api/v1/admin/ledger/export.xlsx?unit_id={first_unit_id}",
        headers=headers,
    )
    assert current.status_code == 200, current.text
    workbook = load_workbook(BytesIO(current.content), data_only=True)
    assert workbook.sheetnames == ["订单台账", "商品需求汇总"]
    detail = worksheet_rows(workbook["订单台账"])
    assert detail[0] == (
        "序号", "订单编号", "商品分类", "商品名称", "规格", "计量单位", "数量", "单价", "小计", "订单金额", "订单状态", "下单时间", "单位名称"
    )
    assert {row[1] for row in detail[1:]} == {first["order_no"], second["order_no"]}
    archived_completed = next(row for row in detail[1:] if row[1] == first["order_no"])
    assert archived_completed[7:12] == (3.21, 9.63, 9.63, "已完成", "2026-08-01 23:59:59")
    summary = worksheet_rows(workbook["商品需求汇总"])
    assert summary[0] == ("商品分类", "商品名称", "规格", "计量单位", "需求数量", "需求金额")
    assert {(row[1], row[3], row[4]) for row in summary[1:]} == {("土豆", "斤", 3), ("土豆", "箱", 2)}

    current_day = client.get(
        "/api/v1/admin/ledger/export.xlsx?start_date=2026-08-02&end_date=2026-08-02",
        headers=headers,
    )
    assert current_day.status_code == 200
    current_day_book = load_workbook(BytesIO(current_day.content), data_only=True)
    assert {row[1] for row in worksheet_rows(current_day_book["订单台账"])[1:]} == {second["order_no"], third["order_no"]}

    all_history = client.get(
        f"/api/v1/admin/ledger/export.xlsx?all=true&unit_id={first_unit_id}",
        headers=headers,
    )
    assert all_history.status_code == 200, all_history.text
    all_book = load_workbook(BytesIO(all_history.content), data_only=True)
    assert {row[1] for row in worksheet_rows(all_book["订单台账"])[1:]} == {first["order_no"], second["order_no"], third["order_no"]}
    assert "全部" in unquote(all_history.headers["content-disposition"])

    unfiltered = client.get("/api/v1/admin/ledger/export.xlsx", headers=headers)
    assert unfiltered.status_code == 200, unfiltered.text
    assert "全部" in unquote(unfiltered.headers["content-disposition"])

    empty_export = client.get("/api/v1/admin/ledger/export.xlsx?unit_id=does-not-exist", headers=headers)
    assert empty_export.status_code == 400
    assert empty_export.json()["detail"] == "当前筛选条件下没有可导出的采购台账"


def test_phase5_ledger_thousand_order_query_and_export_stay_single_query_path(tmp_path):
    client = make_client(tmp_path)
    headers = login(client, "root_admin", "StrongPassword123")
    product = create_product(client, headers, "LEDGER-PERF", "性能土豆", "斤", 250, "蔬菜")
    unit = client.post(
        "/api/v1/admin/units",
        headers=headers,
        json={"unit_code": "LEDGER-PERF", "unit_name": "性能测试单位", "default_delivery_point": "性能测试收货点"},
    ).json()

    from app.database import connect

    with connect() as conn:
        orders = []
        items = []
        for index in range(1000):
            order_id = str(uuid4())
            orders.append((order_id, f"SPPERF{index:06d}", unit["id"], unit["unit_name"], unit["default_delivery_point"], "completed", 500, "2026-08-10 04:00:00"))
            items.append((str(uuid4()), order_id, product["id"], product["product_code"], product["name"], product["category"], product["spec"], product["unit"], 250, "2", 500))
        conn.executemany(
            """
            INSERT INTO orders(id, order_no, unit_id, unit_name_snapshot, delivery_point_snapshot, status, total_cents, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            orders,
        )
        conn.executemany(
            """
            INSERT INTO order_items(id, order_id, product_id, product_code_snapshot, product_name_snapshot, category_snapshot, spec_snapshot, unit_snapshot, price_cents_snapshot, quantity, subtotal_cents)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            items,
        )
        conn.commit()

    query_started = perf_counter()
    page = client.get("/api/v1/admin/ledger?page=1&page_size=20", headers=headers)
    query_seconds = perf_counter() - query_started
    assert page.status_code == 200, page.text
    assert page.json()["total"] == 1000
    assert query_seconds < 10

    export_started = perf_counter()
    exported = client.get("/api/v1/admin/ledger/export.xlsx", headers=headers)
    export_seconds = perf_counter() - export_started
    assert exported.status_code == 200, exported.text
    assert export_seconds < 10
    workbook = load_workbook(BytesIO(exported.content), read_only=True, data_only=True)
    assert workbook["订单台账"].max_row == 1001

    with connect() as conn:
        assert conn.execute("PRAGMA integrity_check").fetchone()[0] == "ok"
        assert conn.execute("PRAGMA quick_check").fetchone()[0] == "ok"
        assert conn.execute("PRAGMA foreign_key_check").fetchall() == []
        stock = conn.execute("SELECT COUNT(*) FROM products WHERE CAST(stock_quantity AS REAL) < 0").fetchone()[0]
        reserved = conn.execute("SELECT COUNT(*) FROM products WHERE CAST(reserved_quantity AS REAL) < 0").fetchone()[0]
    assert stock == 0
    assert reserved == 0
