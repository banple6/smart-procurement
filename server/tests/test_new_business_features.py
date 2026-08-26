import os
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook


def make_client(tmp_path):
    os.environ["APP_ENV"] = "test"
    os.environ["APP_SECRET"] = "test-secret"
    os.environ["DATABASE_PATH"] = str(tmp_path / "smart_procurement.db")
    os.environ["UPLOAD_DIR"] = str(tmp_path / "uploads")
    os.environ["PRIVATE_UPLOAD_DIR"] = str(tmp_path / "private_uploads")
    os.environ["INITIAL_ADMIN_USERNAME"] = "root_admin"
    os.environ["INITIAL_ADMIN_PASSWORD"] = "StrongPassword123"

    from app.database import init_db
    from app.main import app, seed_initial_admin

    init_db()
    seed_initial_admin()
    return TestClient(app)


def login(client, username, password):
    response = client.post("/api/v1/auth/login", json={"username": username, "password": password})
    assert response.status_code == 200, response.text
    return {"Authorization": f"Bearer {response.json()['token']}"}


def create_product(
    client,
    headers,
    code="VEG-001",
    name="土豆",
    unit="斤",
    price_cents=200,
    category="蔬菜",
):
    response = client.post(
        "/api/v1/admin/products",
        headers=headers,
        json={
            "product_code": code,
            "name": name,
            "category": category,
            "spec": "散装" if unit in {"公斤", "斤"} else "预包装",
            "unit": unit,
            "price_cents": price_cents,
            "stock_quantity": "1000",
            "min_order_quantity": "1",
            "quantity_step": "1",
            "supply_status": "normal",
            "active": True,
        },
    )
    assert response.status_code == 200, response.text
    return response.json()


def create_unit_order(client, admin_headers, product_id, suffix, quantity):
    unit = client.post(
        "/api/v1/admin/units",
        headers=admin_headers,
        json={
            "unit_code": f"U-{suffix}",
            "unit_name": f"第{suffix}子单位",
            "default_delivery_point": f"第{suffix}收货点",
        },
    )
    assert unit.status_code == 200, unit.text
    username = f"unit_{suffix}"
    password = f"Unit{suffix}Password123"
    user = client.post(
        "/api/v1/admin/users",
        headers=admin_headers,
        json={
            "username": username,
            "password": password,
            "display_name": f"第{suffix}账号",
            "role": "unit_user",
            "unit_id": unit.json()["id"],
            "must_change_password": False,
        },
    )
    assert user.status_code == 200, user.text
    unit_headers = login(client, username, password)
    order = client.post(
        "/api/v1/orders",
        headers=unit_headers,
        json={"items": [{"product_id": product_id, "quantity": str(quantity)}]},
    )
    assert order.status_code == 200, order.text
    return order.json()


def advance_order_to_preparing(client, admin_headers, order):
    response = client.patch(
        f"/api/v1/admin/orders/{order['id']}/status",
        headers=admin_headers,
        json={"status": "accepted"},
    )
    assert response.status_code == 200, response.text
    assert response.json()["status"] == "preparing"
    return response.json()


def test_0020_and_0021_migrations_create_explicit_batch_and_soft_delete_columns(tmp_path):
    make_client(tmp_path)
    from app.database import connect, migration_status

    status = migration_status()
    assert status["pending"] == []
    assert status["applied"][-3:] == [
        "0020_delivery_batches",
        "0021_order_soft_delete",
        "0022_accept_immediately_preparing",
    ]
    with connect() as conn:
        tables = {row["name"] for row in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")}
        assert {"delivery_batches", "delivery_batch_orders"} <= tables
        order_columns = {row["name"] for row in conn.execute("PRAGMA table_info(orders)")}
        assert {"is_deleted", "deleted_at", "deleted_by"} <= order_columns


def test_0022_migration_moves_legacy_accepted_orders_to_preparing(tmp_path):
    client = make_client(tmp_path)
    headers = login(client, "root_admin", "StrongPassword123")
    product = create_product(client, headers)
    order = create_unit_order(client, headers, product["id"], "MIG", "2")

    from app.database import apply_accept_immediately_preparing_migration, connect

    with connect() as conn:
        conn.execute(
            "UPDATE orders SET status = 'accepted', accepted_at = NULL, preparing_at = NULL WHERE id = ?",
            (order["id"],),
        )
        apply_accept_immediately_preparing_migration(conn)
        migrated = conn.execute(
            "SELECT status, accepted_at, preparing_at, version FROM orders WHERE id = ?",
            (order["id"],),
        ).fetchone()
    assert migrated["status"] == "preparing"
    assert migrated["accepted_at"]
    assert migrated["preparing_at"]
    assert migrated["version"] == order["version"] + 1


def test_standard_template_and_product_batch_archive_keep_product_rows(tmp_path):
    client = make_client(tmp_path)
    headers = login(client, "root_admin", "StrongPassword123")
    first = create_product(client, headers, "VEG-001", "土豆")
    second = create_product(client, headers, "VEG-002", "大白菜")

    template = client.get("/api/v1/admin/products/import-template.xlsx", headers=headers)
    assert template.status_code == 200, template.text
    workbook = load_workbook(BytesIO(template.content), data_only=True)
    assert workbook.sheetnames == ["食材导入", "填写说明"]
    assert [cell.value for cell in workbook["食材导入"][1]] == ["商品编码", "商品名称", "商品分类", "规格", "计量单位", "单价（元）", "库存"]

    archived = client.request(
        "DELETE",
        "/api/v1/admin/products/batch",
        headers=headers,
        json={"ids": [first["id"], second["id"]], "confirmed": True},
    )
    assert archived.status_code == 200, archived.text
    assert archived.json()["archived_count"] == 2
    repeated = client.request(
        "DELETE",
        "/api/v1/admin/products/batch",
        headers=headers,
        json={"ids": [first["id"], second["id"]], "confirmed": True},
    )
    assert repeated.status_code == 200
    assert repeated.json()["already_archived_count"] == 2

    from app.database import connect

    with connect() as conn:
        rows = conn.execute("SELECT id, is_deleted, active FROM products ORDER BY id").fetchall()
        assert len(rows) == 2
        assert all(row["is_deleted"] == 1 and row["active"] == 0 for row in rows)


def test_clear_product_catalog_requires_exact_count_and_preserves_order_history(tmp_path):
    client = make_client(tmp_path)
    headers = login(client, "root_admin", "StrongPassword123")
    product = create_product(client, headers)
    order = create_unit_order(client, headers, product["id"], "CAT", "2")

    stale = client.request(
        "DELETE",
        "/api/v1/admin/products/all",
        headers=headers,
        json={"confirmed": True, "confirmation_text": "确认删除", "expected_count": 0},
    )
    assert stale.status_code == 409, stale.text

    cleared = client.request(
        "DELETE",
        "/api/v1/admin/products/all",
        headers=headers,
        json={"confirmed": True, "confirmation_text": "确认删除", "expected_count": 1},
    )
    assert cleared.status_code == 200, cleared.text
    assert cleared.json()["archived_count"] == 1
    assert client.get("/api/v1/admin/products", headers=headers).json() == []

    from app.database import connect

    with connect() as conn:
        stored_product = conn.execute("SELECT is_deleted, active FROM products WHERE id = ?", (product["id"],)).fetchone()
        stored_item = conn.execute(
            "SELECT product_name_snapshot, price_cents_snapshot, subtotal_cents FROM order_items WHERE order_id = ?",
            (order["id"],),
        ).fetchone()
    assert stored_product["is_deleted"] == 1
    assert stored_product["active"] == 0
    assert stored_item["product_name_snapshot"] == "土豆"
    assert stored_item["price_cents_snapshot"] == 200
    assert stored_item["subtotal_cents"] == 400


def test_delivery_batch_membership_and_summary_use_one_explicit_scope(tmp_path):
    client = make_client(tmp_path)
    admin_headers = login(client, "root_admin", "StrongPassword123")
    product = create_product(client, admin_headers)
    first = create_unit_order(client, admin_headers, product["id"], "01", "100")
    second = create_unit_order(client, admin_headers, product["id"], "02", "200")
    advance_order_to_preparing(client, admin_headers, first)
    advance_order_to_preparing(client, admin_headers, second)
    batch = client.post(
        "/api/v1/admin/batches",
        headers=admin_headers,
        json={"name": "第一配送批次", "order_ids": [first["id"], second["id"]]},
    )
    assert batch.status_code == 200, batch.text
    assert {order["id"] for order in batch.json()["orders"]} == {first["id"], second["id"]}

    summary = client.get(f"/api/v1/admin/batches/{batch.json()['id']}/summary", headers=admin_headers)
    assert summary.status_code == 200, summary.text
    payload = summary.json()
    assert payload["unit_count"] == 2
    assert payload["order_count"] == 2
    assert [(unit["unit_name"], unit["items"][0]["quantity"]) for unit in payload["by_unit"]] == [
        ("第01子单位", "100"),
        ("第02子单位", "200"),
    ]
    assert payload["by_product"][0]["product_name"] == "土豆"
    assert payload["by_product"][0]["total_quantity"] == "300"
    assert [item["quantity"] for item in payload["by_product"][0]["unit_breakdown"]] == ["100", "200"]

    picking = client.get(f"/api/v1/admin/batches/{batch.json()['id']}/picking-list.xlsx", headers=admin_headers)
    assert picking.status_code == 200, picking.text
    workbook = load_workbook(BytesIO(picking.content), data_only=True)
    assert "总计" in workbook.sheetnames
    assert "蔬菜" in workbook.sheetnames
    assert workbook["总计"][4][1].value == "土豆"
    assert str(workbook["总计"][4][4].value) == "300"
    assert workbook["总计"].cell(6, 1).value.startswith("接收人签字")

    eligible = client.get("/api/v1/admin/batches/eligible-orders", headers=admin_headers)
    assert eligible.status_code == 200, eligible.text
    assert {item["id"] for item in eligible.json()["items"]}.isdisjoint({first["id"], second["id"]})


def test_delivery_batch_invalid_ids_return_chinese_404(tmp_path):
    client = make_client(tmp_path)
    headers = login(client, "root_admin", "StrongPassword123")
    for suffix in ("summary", "summary.xlsx", "picking-list.xlsx", "outbound.xlsx"):
        response = client.get(f"/api/v1/admin/batches/missing-batch/{suffix}", headers=headers)
        assert response.status_code == 404, response.text
        assert response.json()["detail"] == "配送批次不存在"


def test_delivery_batch_aggregates_sixteen_units_without_n_plus_one_semantics(tmp_path):
    client = make_client(tmp_path)
    headers = login(client, "root_admin", "StrongPassword123")
    product = create_product(client, headers)
    orders = [
        create_unit_order(client, headers, product["id"], f"{index:02d}", str(index))
        for index in range(1, 17)
    ]
    for order in orders:
        advance_order_to_preparing(client, headers, order)
    batch = client.post(
        "/api/v1/admin/batches",
        headers=headers,
        json={"name": "十六单位配送批次", "order_ids": [order["id"] for order in orders]},
    )
    assert batch.status_code == 200, batch.text
    summary = client.get(f"/api/v1/admin/batches/{batch.json()['id']}/summary", headers=headers)
    assert summary.status_code == 200, summary.text
    payload = summary.json()
    assert payload["unit_count"] == 16
    assert payload["order_count"] == 16
    assert len(payload["by_unit"]) == 16
    assert payload["by_product"][0]["total_quantity"] == "136"
    assert len(payload["by_product"][0]["unit_breakdown"]) == 16


def test_batch_summary_keeps_incompatible_units_separate(tmp_path):
    client = make_client(tmp_path)
    headers = login(client, "root_admin", "StrongPassword123")
    product = create_product(client, headers, unit="斤")
    first = create_unit_order(client, headers, product["id"], "UNITA", "10")
    advance_order_to_preparing(client, headers, first)
    current = client.get(f"/api/v1/products/{product['id']}", headers=headers).json()
    changed = client.put(
        f"/api/v1/admin/products/{product['id']}",
        headers=headers,
        json={"unit": "公斤", "expected_version": current["version"]},
    )
    assert changed.status_code == 200, changed.text
    second = create_unit_order(client, headers, product["id"], "UNITB", "5")
    advance_order_to_preparing(client, headers, second)
    batch = client.post(
        "/api/v1/admin/batches",
        headers=headers,
        json={"name": "不同单位批次", "order_ids": [first["id"], second["id"]]},
    )
    assert batch.status_code == 200, batch.text
    summary = client.get(f"/api/v1/admin/batches/{batch.json()['id']}/summary", headers=headers).json()
    assert [(item["unit"], item["total_quantity"]) for item in summary["by_product"]] == [
        ("斤", "10"),
        ("公斤", "5"),
    ]


def test_accept_immediately_enters_preparing_and_repeated_accept_is_idempotent(tmp_path):
    client = make_client(tmp_path)
    admin_headers = login(client, "root_admin", "StrongPassword123")
    product = create_product(client, admin_headers)
    order = create_unit_order(client, admin_headers, product["id"], "03", "5")

    accepted = client.patch(
        f"/api/v1/admin/orders/{order['id']}/status",
        headers=admin_headers,
        json={"status": "accepted", "expected_status": "pending", "expected_version": order["version"]},
    )
    assert accepted.status_code == 200, accepted.text
    first_payload = accepted.json()
    assert first_payload["status"] == "preparing"
    assert first_payload["accepted_at"]
    assert first_payload["preparing_at"]

    repeated = client.patch(
        f"/api/v1/admin/orders/{order['id']}/status",
        headers=admin_headers,
        json={"status": "accepted"},
    )
    assert repeated.status_code == 200, repeated.text
    assert repeated.json()["status"] == "preparing"
    assert repeated.json()["version"] == first_payload["version"]

    from app.database import connect

    with connect() as conn:
        logs = conn.execute(
            "SELECT COUNT(*) AS c FROM order_logs WHERE order_id = ? AND action = 'accept'",
            (order["id"],),
        ).fetchone()["c"]
        assert logs == 1


def test_pending_order_delete_archives_releases_inventory_and_preserves_history(tmp_path):
    client = make_client(tmp_path)
    admin_headers = login(client, "root_admin", "StrongPassword123")
    product = create_product(client, admin_headers)
    order = create_unit_order(client, admin_headers, product["id"], "04", "7")
    advance_order_to_preparing(client, admin_headers, order)
    batch = client.post(
        "/api/v1/admin/batches",
        headers=admin_headers,
        json={"name": "待删除订单批次", "order_ids": [order["id"]]},
    )
    assert batch.status_code == 200, batch.text

    voided = client.post(
        f"/api/v1/admin/orders/{order['id']}/void",
        headers=admin_headers,
        json={"reason": "测试作废后归档"},
    )
    assert voided.status_code == 200, voided.text

    deleted = client.request(
        "DELETE",
        f"/api/v1/admin/orders/{order['id']}",
        headers=admin_headers,
        json={"reason": "测试软删除"},
    )
    assert deleted.status_code == 200, deleted.text
    assert deleted.json()["already_deleted"] is False
    repeated = client.request(
        "DELETE",
        f"/api/v1/admin/orders/{order['id']}",
        headers=admin_headers,
        json={"reason": "重复删除"},
    )
    assert repeated.status_code == 200, repeated.text
    assert repeated.json()["already_deleted"] is True

    active_orders = client.get("/api/v1/admin/orders", headers=admin_headers)
    assert active_orders.status_code == 200
    assert all(item["id"] != order["id"] for item in active_orders.json()["items"])
    archived_orders = client.get("/api/v1/admin/orders?archived=true", headers=admin_headers)
    assert archived_orders.status_code == 200
    assert any(item["id"] == order["id"] for item in archived_orders.json()["items"])
    assert client.get(f"/api/v1/admin/orders/{order['id']}", headers=admin_headers).status_code == 200

    ledger = client.get("/api/v1/admin/ledger", headers=admin_headers)
    assert ledger.status_code == 200
    assert any(item["order_id"] == order["id"] for item in ledger.json())
    summary = client.get(f"/api/v1/admin/batches/{batch.json()['id']}/summary", headers=admin_headers)
    assert summary.status_code == 200
    assert summary.json()["order_count"] == 0

    from app.database import connect

    with connect() as conn:
        stored = conn.execute("SELECT status, is_deleted, archived_at FROM orders WHERE id = ?", (order["id"],)).fetchone()
        product_row = conn.execute("SELECT reserved_quantity FROM products WHERE id = ?", (product["id"],)).fetchone()
        soft_delete_logs = conn.execute(
            "SELECT COUNT(*) AS c FROM order_logs WHERE order_id = ? AND action = 'soft_delete'",
            (order["id"],),
        ).fetchone()["c"]
        assert stored["status"] == "voided"
        assert stored["is_deleted"] == 0
        assert stored["archived_at"]
        batch_link = conn.execute(
            "SELECT COUNT(*) AS c FROM delivery_batch_orders WHERE batch_id = ? AND order_id = ?",
            (batch.json()["id"], order["id"]),
        ).fetchone()["c"]
        assert str(product_row["reserved_quantity"]) in {"0", "0.0"}
        assert soft_delete_logs == 1
        assert batch_link == 1


def test_non_terminal_delete_is_blocked_and_fulfilled_orders_are_archived(tmp_path):
    client = make_client(tmp_path)
    headers = login(client, "root_admin", "StrongPassword123")
    product = create_product(client, headers)
    preparing_order = create_unit_order(client, headers, product["id"], "DELA", "1")
    accepted = client.patch(
        f"/api/v1/admin/orders/{preparing_order['id']}/status",
        headers=headers,
        json={"status": "accepted"},
    )
    assert accepted.status_code == 200, accepted.text
    blocked = client.request(
        "DELETE",
        f"/api/v1/admin/orders/{preparing_order['id']}",
        headers=headers,
        json={"reason": "不应直接删除"},
    )
    assert blocked.status_code == 409
    assert blocked.json()["detail"] == "请先作废订单，再执行删除"

    fulfilled_order = create_unit_order(client, headers, product["id"], "DELB", "1")
    from app.database import connect

    with connect() as conn:
        conn.execute(
            "UPDATE orders SET status = 'completed', completed_at = CURRENT_TIMESTAMP WHERE id = ?",
            (fulfilled_order["id"],),
        )
    completed_delete = client.request(
        "DELETE",
        f"/api/v1/admin/orders/{fulfilled_order['id']}",
        headers=headers,
        json={"reason": "不应删除已完成订单"},
    )
    assert completed_delete.status_code == 200, completed_delete.text
    assert completed_delete.json()["archived"] is True
    archived = client.get("/api/v1/admin/orders?archived=true", headers=headers)
    assert any(row["id"] == fulfilled_order["id"] for row in archived.json()["items"])

    with connect() as conn:
        stored = conn.execute("SELECT status, is_deleted, archived_at FROM orders WHERE id = ?", (fulfilled_order["id"],)).fetchone()
        item_count = conn.execute("SELECT COUNT(*) AS c FROM order_items WHERE order_id = ?", (fulfilled_order["id"],)).fetchone()["c"]
    assert stored["status"] == "completed"
    assert stored["is_deleted"] == 0
    assert stored["archived_at"]
    assert item_count == 1


def test_shipped_order_delete_archives_without_removing_business_history(tmp_path):
    client = make_client(tmp_path)
    headers = login(client, "root_admin", "StrongPassword123")
    product = create_product(client, headers)
    order = create_unit_order(client, headers, product["id"], "DELS", "1")
    from app.database import connect

    with connect() as conn:
        conn.execute(
            "UPDATE orders SET status = 'shipped', shipped_at = CURRENT_TIMESTAMP WHERE id = ?",
            (order["id"],),
        )
        conn.commit()

    deleted = client.request(
        "DELETE",
        f"/api/v1/admin/orders/{order['id']}",
        headers=headers,
        json={"reason": "已发货订单归档"},
    )
    assert deleted.status_code == 200, deleted.text
    assert deleted.json()["archived"] is True

    with connect() as conn:
        stored = conn.execute("SELECT status, is_deleted, archived_at FROM orders WHERE id = ?", (order["id"],)).fetchone()
        item = conn.execute(
            "SELECT price_cents_snapshot, subtotal_cents FROM order_items WHERE order_id = ?",
            (order["id"],),
        ).fetchone()
    assert stored["status"] == "shipped"
    assert stored["is_deleted"] == 0
    assert stored["archived_at"]
    assert item["price_cents_snapshot"] > 0
    assert item["subtotal_cents"] > 0


def test_product_spec_tracks_unit_defaults_without_overwriting_custom_spec(tmp_path):
    client = make_client(tmp_path)
    headers = login(client, "root_admin", "StrongPassword123")
    product = create_product(client, headers, code="PACK-001", name="瓶装水", unit="箱")
    assert product["spec"] == "预包装"

    changed_to_bulk = client.put(
        f"/api/v1/admin/products/{product['id']}",
        headers=headers,
        json={"unit": "公斤", "expected_version": product["version"]},
    )
    assert changed_to_bulk.status_code == 200, changed_to_bulk.text
    assert changed_to_bulk.json()["spec"] == "散装"

    custom = client.put(
        f"/api/v1/admin/products/{product['id']}",
        headers=headers,
        json={"unit": "箱", "spec": "12瓶/箱", "expected_version": changed_to_bulk.json()["version"]},
    )
    assert custom.status_code == 200, custom.text
    assert custom.json()["spec"] == "12瓶/箱"

    preserved = client.put(
        f"/api/v1/admin/products/{product['id']}",
        headers=headers,
        json={"unit": "瓶", "expected_version": custom.json()["version"]},
    )
    assert preserved.status_code == 200, preserved.text
    assert preserved.json()["spec"] == "12瓶/箱"


def test_ledger_export_uses_chinese_status_and_order_price_snapshot(tmp_path):
    client = make_client(tmp_path)
    headers = login(client, "root_admin", "StrongPassword123")
    product = create_product(client, headers, price_cents=200)
    create_unit_order(client, headers, product["id"], "05", "3")
    current_product = client.get(f"/api/v1/products/{product['id']}", headers=headers)
    assert current_product.status_code == 200, current_product.text
    changed = client.patch(
        f"/api/v1/admin/products/{product['id']}/price",
        headers=headers,
        json={"price_cents": 999, "expected_version": current_product.json()["version"]},
    )
    assert changed.status_code == 200, changed.text

    exported = client.get("/api/v1/admin/ledger/export.xlsx", headers=headers)
    assert exported.status_code == 200, exported.text
    workbook = load_workbook(BytesIO(exported.content), data_only=True)
    sheet = workbook["订单台账"]
    assert [cell.value for cell in sheet[1]] == [
        "序号", "商品分类", "商品名称", "计量单位", "数量", "单价", "小计", "订单状态", "下单时间", "单位名称"
    ]
    assert sheet.cell(2, 6).value == 2
    assert sheet.cell(2, 7).value == 6
    assert sheet.cell(2, 8).value == "待接单"
    assert [cell.value for cell in workbook["商品需求汇总"][1]] == ["商品分类", "商品名称", "计量单位", "需求数量", "需求金额"]


def test_outbound_workbook_uses_order_snapshot_after_current_price_changes(tmp_path):
    client = make_client(tmp_path)
    headers = login(client, "root_admin", "StrongPassword123")
    product = create_product(client, headers, price_cents=200)
    order = create_unit_order(client, headers, product["id"], "OUT", "3")
    advance_order_to_preparing(client, headers, order)
    batch = client.post(
        "/api/v1/admin/batches",
        headers=headers,
        json={"name": "价格快照批次", "order_ids": [order["id"]]},
    )
    assert batch.status_code == 200, batch.text

    current = client.get(f"/api/v1/products/{product['id']}", headers=headers).json()
    changed = client.patch(
        f"/api/v1/admin/products/{product['id']}/price",
        headers=headers,
        json={"price_cents": 999, "expected_version": current["version"]},
    )
    assert changed.status_code == 200, changed.text

    from app.database import connect

    with connect() as conn:
        conn.execute(
            "UPDATE orders SET status = 'shipped', accepted_at = CURRENT_TIMESTAMP, preparing_at = CURRENT_TIMESTAMP, shipped_at = CURRENT_TIMESTAMP WHERE id = ?",
            (order["id"],),
        )
    outbound = client.get(f"/api/v1/admin/batches/{batch.json()['id']}/outbound.xlsx", headers=headers)
    assert outbound.status_code == 200, outbound.text
    workbook = load_workbook(BytesIO(outbound.content), data_only=True)
    assert workbook.sheetnames == ["总计", "蔬菜"]
    total = workbook["总计"]
    assert [cell.value for cell in total[3]] == ["序号", "商品分类", "商品名称", "计量单位", "需求数量", "单价（元）", "小计（元）"]
    assert total.cell(4, 6).value == 2
    assert total.cell(4, 7).value == 6
    assert total.cell(5, 7).value == 6
    assert total.cell(7, 1).value.startswith("发货人签字")
    assert total.cell(7, 3).value.startswith("接收人签字")
    assert total.cell(7, 6).value.startswith("日期")


def test_isolated_database_integrity_after_new_business_operations(tmp_path):
    client = make_client(tmp_path)
    headers = login(client, "root_admin", "StrongPassword123")
    product = create_product(client, headers)
    order = create_unit_order(client, headers, product["id"], "CHK", "4")
    advance_order_to_preparing(client, headers, order)
    batch = client.post(
        "/api/v1/admin/batches",
        headers=headers,
        json={"name": "完整性检查批次", "order_ids": [order["id"]]},
    )
    assert batch.status_code == 200, batch.text

    from app.database import connect

    with connect() as conn:
        assert conn.execute("PRAGMA integrity_check").fetchone()[0] == "ok"
        assert conn.execute("PRAGMA quick_check").fetchone()[0] == "ok"
        assert conn.execute("PRAGMA foreign_key_check").fetchall() == []
