import os
import sqlite3
import base64
import json
import secrets
import zipfile
from io import BytesIO
from pathlib import Path
from uuid import uuid4

from fastapi.testclient import TestClient
from PIL import Image


def make_client(tmp_path):
    os.environ["APP_ENV"] = "test"
    os.environ["APP_SECRET"] = "test-secret"
    os.environ["DATABASE_PATH"] = str(tmp_path / "smart_procurement.db")
    os.environ["UPLOAD_DIR"] = str(tmp_path / "uploads")
    os.environ["PRIVATE_UPLOAD_DIR"] = str(tmp_path / "private_uploads")
    os.environ["INITIAL_ADMIN_USERNAME"] = "root_admin"
    os.environ["INITIAL_ADMIN_PASSWORD"] = "StrongPassword123"

    from app.main import app, seed_initial_admin
    from app.database import init_db

    init_db()
    seed_initial_admin()
    return TestClient(app)


def login(client, username, password):
    response = client.post("/api/v1/auth/login", json={"username": username, "password": password})
    assert response.status_code == 200, response.text
    token = response.json()["token"]
    return {"Authorization": f"Bearer {token}"}


def set_web_session(client, user_id: str, role: str = "admin", unit_id: str | None = None):
    from app.database import connect
    from app.security import hash_token
    from app.web_session import CSRF_COOKIE, web_session_cookie_name

    token = secrets.token_urlsafe(32)
    csrf = secrets.token_urlsafe(24)
    with connect() as conn:
        conn.execute("UPDATE users SET must_change_password = 0 WHERE id = ?", (user_id,))
        conn.execute(
            """
            INSERT INTO web_sessions(
              id, token_hash, user_id, role, unit_id, idle_expires_at, absolute_expires_at,
              browser_name, browser_os, browser_ip
            )
            VALUES (?, ?, ?, ?, ?, datetime('now', '+30 minutes'), datetime('now', '+8 hours'), 'TestBrowser', 'macOS', '127.0.0.1')
            """,
            (str(uuid4()), hash_token(token), user_id, role, unit_id),
        )
        conn.commit()
    client.cookies.set(web_session_cookie_name(), token)
    client.cookies.set(CSRF_COOKIE, csrf)
    return {"X-CSRF-Token": csrf}


def create_unit_user_product_order(client):
    admin_headers = login(client, "root_admin", "StrongPassword123")
    unit = client.post(
        "/api/v1/admin/units",
        headers=admin_headers,
        json={
            "unit_code": "U001",
            "unit_name": "第一食堂",
            "default_delivery_point": "第一食堂收货点",
        },
    )
    assert unit.status_code == 200, unit.text
    unit_id = unit.json()["id"]

    user = client.post(
        "/api/v1/admin/users",
        headers=admin_headers,
        json={
            "username": "unit001",
            "password": "UnitPassword123",
            "display_name": "第一食堂账号",
            "role": "unit_user",
            "unit_id": unit_id,
            "must_change_password": False,
        },
    )
    assert user.status_code == 200, user.text

    product = client.post(
        "/api/v1/admin/products",
        headers=admin_headers,
        json={
            "product_code": "VEG-TOMATO",
            "name": "西红柿",
            "category": "蔬菜",
            "spec": "普通大红款",
            "unit": "公斤",
            "price_cents": 450,
            "stock_quantity": "10",
            "min_order_quantity": "1",
            "quantity_step": "0.5",
            "warning_quantity": "2",
            "supply_status": "normal",
        },
    )
    assert product.status_code == 200, product.text
    return admin_headers, login(client, "unit001", "UnitPassword123"), unit_id, product.json()["id"]


def test_registration_safety_migration_upgrades_legacy_invite_table(tmp_path):
    os.environ["APP_ENV"] = "test"
    os.environ["APP_SECRET"] = "test-secret"
    os.environ["DATABASE_PATH"] = str(tmp_path / "smart_procurement.db")
    os.environ["UPLOAD_DIR"] = str(tmp_path / "uploads")
    os.environ["PRIVATE_UPLOAD_DIR"] = str(tmp_path / "private_uploads")

    from app.database import column_names, connect, migrate

    migrate()
    with connect() as conn:
        conn.execute("DROP TABLE IF EXISTS manager_registration_requests")
        conn.execute("DROP TABLE IF EXISTS registration_invites")
        conn.execute(
            """
            CREATE TABLE registration_invites (
              id TEXT PRIMARY KEY,
              token_hash TEXT NOT NULL UNIQUE,
              token_suffix TEXT NOT NULL DEFAULT '',
              invite_type TEXT NOT NULL,
              unit_id TEXT,
              max_uses INTEGER NOT NULL DEFAULT 1,
              used_count INTEGER NOT NULL DEFAULT 0,
              phone_required INTEGER NOT NULL DEFAULT 0,
              allowed_phone_hash TEXT NOT NULL DEFAULT '',
              approval_required INTEGER NOT NULL DEFAULT 0,
              active INTEGER NOT NULL DEFAULT 1,
              expires_at TEXT NOT NULL,
              created_by TEXT,
              created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
              revoked_at TEXT,
              revoked_by TEXT,
              note TEXT NOT NULL DEFAULT ''
            )
            """
        )
        conn.execute("DELETE FROM schema_migrations WHERE version IN ('0008_registration_safety', '0009_manager_registration_requests')")
        conn.execute(
            """
            INSERT INTO registration_invites(
              id, token_hash, token_suffix, invite_type, unit_id, max_uses, used_count, active, expires_at
            )
            VALUES
              ('legacy-unit', 'hash-unit', 'U123', 'unit', 'unit-1', 2, 0, 1, datetime('now', '+1 day')),
              ('legacy-manager', 'hash-manager', 'M123', 'manager', NULL, 1, 0, 0, datetime('now', '+1 day'))
            """
        )
        conn.commit()

    applied = migrate()
    assert "0008_registration_safety" in applied
    assert "0009_manager_registration_requests" in applied

    with connect() as conn:
        columns = column_names(conn, "registration_invites")
        assert {"display_code_suffix", "role", "allowed_phone_masked", "status", "updated_at"} <= columns
        unit_invite = conn.execute(
            "SELECT display_code_suffix, role, status, updated_at FROM registration_invites WHERE id = 'legacy-unit'"
        ).fetchone()
        manager_invite = conn.execute(
            "SELECT display_code_suffix, role, status FROM registration_invites WHERE id = 'legacy-manager'"
        ).fetchone()
    assert unit_invite["display_code_suffix"] == "U123"
    assert unit_invite["role"] == "unit_user"
    assert unit_invite["status"] == "active"
    assert unit_invite["updated_at"]
    assert manager_invite["display_code_suffix"] == "M123"
    assert manager_invite["role"] == "admin"
    assert manager_invite["status"] == "revoked"


def sample_image_bytes(size=(32, 24), color=(180, 40, 30)) -> bytes:
    image = Image.new("RGB", size, color)
    out = BytesIO()
    image.save(out, format="JPEG")
    return out.getvalue()


def sample_apk_bytes(version_code: int = 10, signer: str = "ABCDEF1234567890ABCDEF1234567890ABCDEF1234567890ABCDEF1234567890") -> BytesIO:
    out = BytesIO()
    with zipfile.ZipFile(out, "w") as archive:
        archive.writestr(
            "META-INF/JRXP-APK-METADATA.json",
            json.dumps(
                {
                    "package_name": "com.smartprocurement.internal",
                    "version_code": version_code,
                    "version_name": f"1.1-test{version_code}",
                    "channel": "staging",
                    "signer_sha256": signer,
                    "min_sdk": 24,
                }
            ),
        )
    out.seek(0)
    return out


def advance_to_preparing(client, admin_headers, order_id):
    accepted = client.patch(f"/api/v1/admin/orders/{order_id}/status", headers=admin_headers, json={"status": "accepted"})
    assert accepted.status_code == 200, accepted.text
    preparing = client.patch(f"/api/v1/admin/orders/{order_id}/status", headers=admin_headers, json={"status": "preparing"})
    assert preparing.status_code == 200, preparing.text
    return preparing.json()


def test_auth_and_disabled_account(tmp_path):
    client = make_client(tmp_path)
    assert client.post("/api/v1/auth/login", json={"username": "root_admin", "password": "bad"}).status_code == 401
    admin_headers = login(client, "root_admin", "StrongPassword123")
    unit = client.post(
        "/api/v1/admin/units",
        headers=admin_headers,
        json={"unit_code": "U002", "unit_name": "第二食堂", "default_delivery_point": "二号点"},
    ).json()
    user = client.post(
        "/api/v1/admin/users",
        headers=admin_headers,
        json={
            "username": "disabled",
            "password": "Disabled123",
            "display_name": "停用账号",
            "role": "unit_user",
            "unit_id": unit["id"],
        },
    ).json()
    stopped = client.patch(f"/api/v1/admin/users/{user['id']}/status", headers=admin_headers, json={"active": False})
    assert stopped.status_code == 200
    assert client.post("/api/v1/auth/login", json={"username": "disabled", "password": "Disabled123"}).status_code == 403


def test_login_and_me_return_unit_profile_fields(tmp_path):
    client = make_client(tmp_path)
    admin_headers, unit_headers, unit_id, _ = create_unit_user_product_order(client)

    login_response = client.post("/api/v1/auth/login", json={"username": "unit001", "password": "UnitPassword123"})
    assert login_response.status_code == 200, login_response.text
    user = login_response.json()["user"]
    assert user["username"] == "unit001"
    assert user["display_name"] == "第一食堂账号"
    assert user["role"] == "unit_user"
    assert user["unit_id"] == unit_id
    assert user["unit_code"] == "U001"
    assert user["unit_name"] == "第一食堂"
    assert user["default_delivery_point"] == "第一食堂收货点"
    assert user["must_change_password"] is False

    me = client.get("/api/v1/auth/me", headers=unit_headers)
    assert me.status_code == 200, me.text
    assert me.json()["unit_code"] == "U001"
    assert me.json()["unit_name"] == "第一食堂"
    assert me.json()["default_delivery_point"] == "第一食堂收货点"

    admin_me = client.get("/api/v1/auth/me", headers=admin_headers)
    assert admin_me.status_code == 200
    assert admin_me.json()["unit_id"] == ""
    assert admin_me.json()["unit_name"] == ""


def test_change_password_requires_complex_password_and_new_login(tmp_path):
    client = make_client(tmp_path)
    _, unit_headers, _, _ = create_unit_user_product_order(client)

    too_short = client.post(
        "/api/v1/auth/change-password",
        headers=unit_headers,
        json={"old_password": "UnitPassword123", "new_password": "abc123"},
    )
    assert too_short.status_code == 400
    assert too_short.json()["detail"] == "密码至少 8 位，且包含字母和数字"

    same_as_username = client.post(
        "/api/v1/auth/change-password",
        headers=unit_headers,
        json={"old_password": "UnitPassword123", "new_password": "unit001"},
    )
    assert same_as_username.status_code == 400
    assert same_as_username.json()["detail"] == "新密码不能与账号相同"

    changed = client.post(
        "/api/v1/auth/change-password",
        headers=unit_headers,
        json={"old_password": "UnitPassword123", "new_password": "NewPass123"},
    )
    assert changed.status_code == 200
    assert client.get("/api/v1/auth/me", headers=unit_headers).status_code == 401
    assert client.post("/api/v1/auth/login", json={"username": "unit001", "password": "UnitPassword123"}).status_code == 401
    assert client.post("/api/v1/auth/login", json={"username": "unit001", "password": "NewPass123"}).status_code == 200


def test_role_permissions_and_unit_order_isolation(tmp_path):
    client = make_client(tmp_path)
    admin_headers, unit_headers, unit_id, product_id = create_unit_user_product_order(client)
    forbidden = client.post(
        "/api/v1/admin/products",
        headers=unit_headers,
        json={"product_code": "X", "name": "X", "category": "其他", "spec": "x", "unit": "个", "price_cents": 1, "stock_quantity": "1"},
    )
    assert forbidden.status_code == 403

    order = client.post(
        "/api/v1/orders",
        headers=unit_headers,
        json={"note": "上午送达", "items": [{"product_id": product_id, "quantity": "2"}]},
    )
    assert order.status_code == 200, order.text
    assert order.json()["unit_id"] == unit_id
    assert order.json()["created_by_username"] == "unit001"
    assert "created_by" not in order.json()
    assert order.json()["total_cents"] == 900
    assert client.get("/api/v1/orders", headers=unit_headers).json()["total"] == 1
    assert client.get("/api/v1/admin/orders", headers=admin_headers).json()["total"] == 1


def test_unit_web_pages_recover_missing_root_csrf_cookie(tmp_path):
    client = make_client(tmp_path)
    _, _, unit_id, product_id = create_unit_user_product_order(client)
    from app.database import connect, one

    with connect() as conn:
        user = one(conn, "SELECT id FROM users WHERE username = ?", ("unit001",))
    set_web_session(client, user["id"], role="unit_user", unit_id=unit_id)
    client.cookies.delete("csrf_token")

    page = client.get("/unit/products")
    assert page.status_code == 200, page.text
    csrf = page.cookies.get("csrf_token")
    assert csrf

    added = client.post(
        "/unit/cart/items",
        headers={"X-CSRF-Token": csrf},
        json={"product_id": product_id, "quantity": "1"},
    )
    assert added.status_code == 200, added.text
    assert added.json()["item_count"] == 1


def test_app_release_stores_signer_digest_lowercase_for_old_clients(tmp_path):
    client = make_client(tmp_path)
    from fastapi import UploadFile
    from app.database import connect, one
    from app.services import app_update

    with connect() as conn:
        admin = one(conn, "SELECT id, role FROM users WHERE username = ?", ("root_admin",))
    release = app_update.create_release(
        admin=admin,
        apk=UploadFile(file=sample_apk_bytes(), filename="app-test.apk"),
        channel="staging",
        package_name="com.smartprocurement.internal",
        update_type="recommended",
        title="测试版本",
        release_notes="签名大小写兼容",
        minimum_supported_version_code=0,
        rollout_percent=100,
    )
    assert release["signer_sha256"] == "abcdef1234567890abcdef1234567890abcdef1234567890abcdef1234567890"


def test_inventory_reservation_price_snapshot_and_status_flow(tmp_path):
    client = make_client(tmp_path)
    admin_headers, unit_headers, _, product_id = create_unit_user_product_order(client)
    order = client.post(
        "/api/v1/orders",
        headers=unit_headers,
        json={"items": [{"product_id": product_id, "quantity": "3"}]},
    ).json()
    product_after_order = client.get(f"/api/v1/products/{product_id}", headers=unit_headers).json()
    assert product_after_order["reserved_quantity"] == "3"
    assert product_after_order["available_quantity"] == "7"

    price_update = client.put(
        f"/api/v1/admin/products/{product_id}",
        headers=admin_headers,
        json={"price_cents": 600, "stock_quantity": "10"},
    )
    assert price_update.status_code == 200
    detail = client.get(f"/api/v1/orders/{order['id']}", headers=unit_headers).json()
    assert detail["items"][0]["price_cents_snapshot"] == 450
    assert detail["items"][0]["subtotal_cents"] == 1350

    too_many = client.post(
        "/api/v1/orders",
        headers=unit_headers,
        json={"items": [{"product_id": product_id, "quantity": "99"}]},
    )
    assert too_many.status_code == 409

    accepted = client.patch(f"/api/v1/admin/orders/{order['id']}/status", headers=admin_headers, json={"status": "accepted"})
    assert accepted.status_code == 200
    assert accepted.json()["accepted_at"]
    assert not accepted.json()["preparing_at"]
    cannot_edit = client.put(
        f"/api/v1/orders/{order['id']}",
        headers=unit_headers,
        json={"items": [{"product_id": product_id, "quantity": "1"}]},
    )
    assert cannot_edit.status_code == 409
    preparing = client.patch(f"/api/v1/admin/orders/{order['id']}/status", headers=admin_headers, json={"status": "preparing"})
    assert preparing.status_code == 200
    assert preparing.json()["accepted_at"]
    assert preparing.json()["preparing_at"]
    assert not preparing.json()["shipped_at"]
    direct_ship = client.patch(f"/api/v1/admin/orders/{order['id']}/status", headers=admin_headers, json={"status": "shipped"})
    assert direct_ship.status_code == 409
    assert direct_ship.json()["detail"] == "请先拍摄并上传发货照片"
    shipped = client.post(
        f"/api/v1/admin/orders/{order['id']}/ship",
        headers=admin_headers,
        data={"note": "已核对数量", "client_request_id": str(uuid4())},
        files=[("photos", ("ship.jpg", sample_image_bytes(), "image/jpeg"))],
    )
    assert shipped.status_code == 200
    assert shipped.json()["preparing_at"]
    assert shipped.json()["shipped_at"]
    assert shipped.json()["shipping_note"] == "已核对数量"
    assert shipped.json()["shipping_photo_count"] == 1
    completed = client.post(f"/api/v1/orders/{order['id']}/confirm-receipt", headers=unit_headers)
    assert completed.status_code == 200
    assert completed.json()["completed_at"]
    completed_product = client.get(f"/api/v1/products/{product_id}", headers=unit_headers).json()
    assert completed_product["stock_quantity"] == "7"
    assert completed_product["reserved_quantity"] == "0"


def test_order_responses_include_unified_status_metadata_and_version(tmp_path):
    client = make_client(tmp_path)
    admin_headers, unit_headers, _, product_id = create_unit_user_product_order(client)
    order = client.post(
        "/api/v1/orders",
        headers=unit_headers,
        json={"items": [{"product_id": product_id, "quantity": "1"}]},
    )
    assert order.status_code == 200, order.text
    body = order.json()
    assert body["status"] == "pending"
    assert body["status_label"] == "待接单"
    assert body["status_stage"] == 1
    assert body["is_terminal"] is False
    assert body["version"] == 1
    assert body["cancelled_at"] == ""

    accepted = client.patch(
        f"/api/v1/admin/orders/{body['id']}/status",
        headers=admin_headers,
        json={"status": "accepted", "expected_status": "pending", "expected_version": 1},
    )
    assert accepted.status_code == 200, accepted.text
    assert accepted.json()["status_label"] == "已接单"
    assert accepted.json()["status_stage"] == 2
    assert accepted.json()["version"] == 2

    stale = client.patch(
        f"/api/v1/admin/orders/{body['id']}/status",
        headers=admin_headers,
        json={"status": "preparing", "expected_status": "pending", "expected_version": 1},
    )
    assert stale.status_code == 409
    assert stale.json()["detail"] == "订单状态已被其他操作员更新，页面已刷新"


def test_cancelled_order_records_cancel_time_and_terminal_status(tmp_path):
    client = make_client(tmp_path)
    _, unit_headers, _, product_id = create_unit_user_product_order(client)
    order = client.post(
        "/api/v1/orders",
        headers=unit_headers,
        json={"items": [{"product_id": product_id, "quantity": "1"}]},
    ).json()

    cancelled = client.post(f"/api/v1/orders/{order['id']}/cancel", headers=unit_headers)
    assert cancelled.status_code == 200, cancelled.text
    body = cancelled.json()
    assert body["status"] == "cancelled"
    assert body["status_label"] == "已取消"
    assert body["status_stage"] == 0
    assert body["is_terminal"] is True
    assert body["cancelled_at"]
    assert body["version"] == 2


def test_ship_order_requires_private_photos_and_returns_authorized_views(tmp_path):
    client = make_client(tmp_path)
    admin_headers, unit_headers, unit_id, product_id = create_unit_user_product_order(client)
    order = client.post(
        "/api/v1/orders",
        headers=unit_headers,
        json={"items": [{"product_id": product_id, "quantity": "1"}]},
    ).json()
    advance_to_preparing(client, admin_headers, order["id"])

    no_photo = client.post(
        f"/api/v1/admin/orders/{order['id']}/ship",
        headers=admin_headers,
        data={"note": "空照片", "client_request_id": str(uuid4())},
    )
    assert no_photo.status_code == 400
    assert no_photo.json()["detail"] == "请至少拍摄一张发货照片"

    too_many = client.post(
        f"/api/v1/admin/orders/{order['id']}/ship",
        headers=admin_headers,
        data={"note": "过多照片", "client_request_id": str(uuid4())},
        files=[("photos", (f"ship{i}.jpg", sample_image_bytes(), "image/jpeg")) for i in range(4)],
    )
    assert too_many.status_code == 400
    assert too_many.json()["detail"] == "最多上传三张发货照片"

    fake = client.post(
        f"/api/v1/admin/orders/{order['id']}/ship",
        headers=admin_headers,
        data={"note": "假图片", "client_request_id": str(uuid4())},
        files=[("photos", ("fake.jpg", b"not-a-real-image", "image/jpeg"))],
    )
    assert fake.status_code == 400
    assert fake.json()["detail"] == "发货照片格式不正确"
    assert client.get(f"/api/v1/orders/{order['id']}", headers=unit_headers).json()["status"] == "preparing"

    request_id = str(uuid4())
    shipped = client.post(
        f"/api/v1/admin/orders/{order['id']}/ship",
        headers=admin_headers,
        data={"note": "已核对数量", "client_request_id": request_id},
        files=[("photos", ("ship.jpg", sample_image_bytes(size=(1800, 1200)), "image/jpeg"))],
    )
    assert shipped.status_code == 200, shipped.text
    body = shipped.json()
    assert body["status"] == "shipped"
    assert body["shipped_at"]
    assert body["shipping_note"] == "已核对数量"
    assert body["shipping_photo_count"] == 1
    assert len(body["shipping_photos"]) == 1
    photo = body["shipping_photos"][0]
    assert photo["thumbnail_url"].startswith(f"/api/v1/orders/{order['id']}/shipping-photos/")
    assert photo["full_url"].endswith("variant=full")

    retry = client.post(
        f"/api/v1/admin/orders/{order['id']}/ship",
        headers=admin_headers,
        data={"note": "重复请求", "client_request_id": request_id},
        files=[("photos", ("ship-again.jpg", sample_image_bytes(color=(10, 40, 90)), "image/jpeg"))],
    )
    assert retry.status_code == 200, retry.text
    assert retry.json()["shipping_photo_count"] == 1

    thumbnail = client.get(photo["thumbnail_url"], headers=unit_headers)
    assert thumbnail.status_code == 200
    assert thumbnail.headers["content-type"] == "image/jpeg"
    assert "no-store" in thumbnail.headers["cache-control"]
    full = client.get(photo["full_url"], headers=admin_headers)
    assert full.status_code == 200
    with Image.open(BytesIO(full.content)) as processed:
        assert processed.format == "JPEG"
        assert max(processed.size) <= 1600
        assert not processed.getexif()

    other_unit = client.post(
        "/api/v1/admin/units",
        headers=admin_headers,
        json={"unit_code": "U-OTHER", "unit_name": "其他单位", "default_delivery_point": "其他收货点"},
    ).json()
    client.post(
        "/api/v1/admin/users",
        headers=admin_headers,
        json={
            "username": "unit_other",
            "password": "OtherPassword123",
            "display_name": "其他单位账号",
            "role": "unit_user",
            "unit_id": other_unit["id"],
            "must_change_password": False,
        },
    )
    other_headers = login(client, "unit_other", "OtherPassword123")
    assert client.get(photo["thumbnail_url"], headers=other_headers).status_code == 404
    assert client.get(photo["thumbnail_url"]).status_code == 401

    from app.database import connect

    with connect() as conn:
        photos = conn.execute("SELECT COUNT(*) AS c FROM order_shipping_photos WHERE order_id = ?", (order["id"],)).fetchone()["c"]
        logs = conn.execute("SELECT COUNT(*) AS c FROM order_logs WHERE order_id = ? AND action = 'ship'", (order["id"],)).fetchone()["c"]
        audits = conn.execute("SELECT COUNT(*) AS c FROM audit_logs WHERE object_id = ? AND action = 'order_ship'", (order["id"],)).fetchone()["c"]
        assert photos == 1
        assert logs == 1
        assert audits == 1


def test_ship_order_rejects_wrong_status_large_files_and_unit_users(tmp_path):
    client = make_client(tmp_path)
    admin_headers, unit_headers, _, product_id = create_unit_user_product_order(client)
    pending_order = client.post(
        "/api/v1/orders",
        headers=unit_headers,
        json={"items": [{"product_id": product_id, "quantity": "1"}]},
    ).json()
    pending_ship = client.post(
        f"/api/v1/admin/orders/{pending_order['id']}/ship",
        headers=admin_headers,
        data={"client_request_id": str(uuid4())},
        files=[("photos", ("ship.jpg", sample_image_bytes(), "image/jpeg"))],
    )
    assert pending_ship.status_code == 409
    assert pending_ship.json()["detail"] == "只有备货中的订单才能确认发货"

    accepted_order = client.post(
        "/api/v1/orders",
        headers=unit_headers,
        json={"items": [{"product_id": product_id, "quantity": "1"}]},
    ).json()
    client.patch(f"/api/v1/admin/orders/{accepted_order['id']}/status", headers=admin_headers, json={"status": "accepted"})
    accepted_ship = client.post(
        f"/api/v1/admin/orders/{accepted_order['id']}/ship",
        headers=admin_headers,
        data={"client_request_id": str(uuid4())},
        files=[("photos", ("ship.jpg", sample_image_bytes(), "image/jpeg"))],
    )
    assert accepted_ship.status_code == 409
    assert accepted_ship.json()["detail"] == "只有备货中的订单才能确认发货"

    preparing_order = client.post(
        "/api/v1/orders",
        headers=unit_headers,
        json={"items": [{"product_id": product_id, "quantity": "1"}]},
    ).json()
    advance_to_preparing(client, admin_headers, preparing_order["id"])
    unit_ship = client.post(
        f"/api/v1/admin/orders/{preparing_order['id']}/ship",
        headers=unit_headers,
        data={"client_request_id": str(uuid4())},
        files=[("photos", ("ship.jpg", sample_image_bytes(), "image/jpeg"))],
    )
    assert unit_ship.status_code == 403

    large = client.post(
        f"/api/v1/admin/orders/{preparing_order['id']}/ship",
        headers=admin_headers,
        data={"client_request_id": str(uuid4())},
        files=[("photos", ("large.jpg", b"x" * (5 * 1024 * 1024 + 1), "image/jpeg"))],
    )
    assert large.status_code == 400
    assert large.json()["detail"] == "发货照片不能超过 5MB"


def test_order_rejects_unavailable_below_min_and_invalid_step(tmp_path):
    client = make_client(tmp_path)
    admin_headers, unit_headers, _, product_id = create_unit_user_product_order(client)

    below_min = client.post(
        "/api/v1/orders",
        headers=unit_headers,
        json={"items": [{"product_id": product_id, "quantity": "0.5"}]},
    )
    assert below_min.status_code == 400
    assert below_min.json()["detail"] == "低于最小申领量"

    wrong_step = client.post(
        "/api/v1/orders",
        headers=unit_headers,
        json={"items": [{"product_id": product_id, "quantity": "1.25"}]},
    )
    assert wrong_step.status_code == 400
    assert wrong_step.json()["detail"] == "数量不符合申领步长"

    zero_stock_product = client.post(
        "/api/v1/admin/products",
        headers=admin_headers,
        json={
            "product_code": "ZERO-STOCK",
            "name": "空库存",
            "category": "蔬菜",
            "spec": "一级",
            "unit": "公斤",
            "price_cents": 300,
            "stock_quantity": "0",
            "min_order_quantity": "1",
            "quantity_step": "1",
            "supply_status": "normal",
        },
    ).json()
    insufficient = client.post(
        "/api/v1/orders",
        headers=unit_headers,
        json={"items": [{"product_id": zero_stock_product["id"], "quantity": "1"}]},
    )
    assert insufficient.status_code == 409
    assert insufficient.json()["detail"] == "库存不足，请减少数量"

    client.patch(f"/api/v1/admin/products/{product_id}/status", headers=admin_headers, json={"supply_status": "paused", "active": True})
    paused = client.post(
        "/api/v1/orders",
        headers=unit_headers,
        json={"items": [{"product_id": product_id, "quantity": "1"}]},
    )
    assert paused.status_code == 409
    assert paused.json()["detail"] == "食材已暂停供应或下架"

    zero_price_product = client.post(
        "/api/v1/admin/products",
        headers=admin_headers,
        json={
            "product_code": "ZERO-PRICE",
            "name": "无价格",
            "category": "蔬菜",
            "spec": "一级",
            "unit": "公斤",
            "price_cents": 0,
            "stock_quantity": "10",
            "min_order_quantity": "1",
            "quantity_step": "1",
            "supply_status": "paused",
            "active": False,
        },
    ).json()
    cannot_enable_zero_price = client.patch(
        f"/api/v1/admin/products/{zero_price_product['id']}/status",
        headers=admin_headers,
        json={"supply_status": "normal", "active": True},
    )
    assert cannot_enable_zero_price.status_code == 400
    assert cannot_enable_zero_price.json()["detail"] == "请先填写商品价格"


def test_product_restore_price_and_inventory_adjustment_are_logged(tmp_path):
    client = make_client(tmp_path)
    admin_headers, unit_headers, _, product_id = create_unit_user_product_order(client)

    assert client.delete(f"/api/v1/admin/products/{product_id}", headers=admin_headers).status_code == 200
    hidden = client.get("/api/v1/products", headers=unit_headers).json()
    assert all(product["id"] != product_id for product in hidden)

    restored = client.post(f"/api/v1/admin/products/{product_id}/restore", headers=admin_headers)
    assert restored.status_code == 200, restored.text
    assert restored.json()["active"] is True
    assert restored.json()["supply_status"] == "normal"

    price = client.patch(f"/api/v1/admin/products/{product_id}/price", headers=admin_headers, json={"price_cents": 575})
    assert price.status_code == 200, price.text
    assert price.json()["price_cents"] == 575

    stock = client.patch(
        f"/api/v1/admin/products/{product_id}/stock",
        headers=admin_headers,
        json={"stock_quantity": "15.5", "detail": "盘点调整"},
    )
    assert stock.status_code == 200, stock.text
    assert stock.json()["stock_quantity"] == "15.5"
    assert stock.json()["available_quantity"] == "15.5"

    from app.database import connect

    with connect() as conn:
        price_logs = conn.execute("SELECT COUNT(*) FROM product_price_logs WHERE product_id = ?", (product_id,)).fetchone()[0]
        stock_logs = conn.execute("SELECT COUNT(*) FROM inventory_logs WHERE product_id = ? AND action = 'admin_adjust'", (product_id,)).fetchone()[0]
    assert price_logs >= 2
    assert stock_logs == 1


def test_cancel_releases_inventory_summary_and_excel(tmp_path):
    client = make_client(tmp_path)
    admin_headers, unit_headers, _, product_id = create_unit_user_product_order(client)
    order = client.post(
        "/api/v1/orders",
        headers=unit_headers,
        json={"items": [{"product_id": product_id, "quantity": "2"}]},
    ).json()
    assert client.post(f"/api/v1/orders/{order['id']}/cancel", headers=unit_headers).status_code == 200
    product = client.get(f"/api/v1/products/{product_id}", headers=unit_headers).json()
    assert product["reserved_quantity"] == "0"
    dashboard = client.get("/api/v1/admin/dashboard", headers=admin_headers)
    assert dashboard.status_code == 200
    summary = client.get("/api/v1/admin/order-summary", headers=admin_headers)
    assert summary.status_code == 200
    ledger = client.get("/api/v1/admin/ledger", headers=admin_headers)
    assert ledger.status_code == 200
    export = client.get("/api/v1/admin/ledger/export.xlsx", headers=admin_headers)
    assert export.status_code == 200
    assert export.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument")
    from urllib.parse import quote
    assert quote("三公鲜配_采购台账_") in export.headers["content-disposition"]
    preparation_export = client.get("/api/v1/admin/preparation-summary/export.xlsx", headers=admin_headers)
    assert preparation_export.status_code == 200
    assert quote("三公鲜配_今日备货单_") in preparation_export.headers["content-disposition"]
    delivery_export = client.get("/api/v1/admin/delivery-sheets/export.xlsx", headers=admin_headers)
    assert delivery_export.status_code == 200
    assert quote("三公鲜配_配送单_") in delivery_export.headers["content-disposition"]
    from io import BytesIO
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(export.content), read_only=True)
    assert workbook.sheetnames == ["订单台账", "商品需求汇总"]
    assert workbook.properties.title == "三公鲜配采购台账"


def test_preparation_summary_includes_current_preparing_orders_from_previous_business_days(tmp_path):
    client = make_client(tmp_path)
    admin_headers, unit_headers, _, product_id = create_unit_user_product_order(client)
    order = client.post(
        "/api/v1/orders",
        headers=unit_headers,
        json={"items": [{"product_id": product_id, "quantity": "2"}]},
    ).json()

    from app.database import connect

    with connect() as conn:
        conn.execute(
            """
            UPDATE orders
            SET status = 'preparing',
                created_at = datetime('now', '-2 days'),
                accepted_at = datetime('now', '-1 day'),
                preparing_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (order["id"],),
        )
        conn.commit()

    current_summary = client.get("/api/v1/admin/preparation-summary", headers=admin_headers)
    assert current_summary.status_code == 200, current_summary.text
    assert current_summary.json()["total"] == 1
    assert current_summary.json()["items"][0]["product_name"] == "西红柿"
    assert current_summary.json()["items"][0]["order_count"] == 1

    today_only = client.get("/api/v1/admin/preparation-summary?business_date=2099-01-01", headers=admin_headers)
    assert today_only.status_code == 200, today_only.text
    assert today_only.json()["total"] == 0


def test_delivery_sheets_include_current_delivery_orders_from_previous_business_days(tmp_path):
    client = make_client(tmp_path)
    admin_headers, unit_headers, _, product_id = create_unit_user_product_order(client)
    order = client.post(
        "/api/v1/orders",
        headers=unit_headers,
        json={"items": [{"product_id": product_id, "quantity": "2"}]},
    ).json()

    from app.database import connect

    with connect() as conn:
        conn.execute(
            """
            UPDATE orders
            SET status = 'preparing',
                created_at = datetime('now', '-2 days'),
                accepted_at = datetime('now', '-1 day'),
                preparing_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (order["id"],),
        )
        conn.commit()

    current_sheets = client.get("/api/v1/admin/delivery-sheets", headers=admin_headers)
    assert current_sheets.status_code == 200, current_sheets.text
    body = current_sheets.json()
    assert body["date_filtered"] is False
    assert len(body["units"]) == 1
    assert body["units"][0]["unit_name"] == "第一食堂"
    assert body["units"][0]["orders"][0]["order_no"] == order["order_no"]
    assert body["units"][0]["orders"][0]["items"][0]["product_name"] == "西红柿"

    future_day = client.get("/api/v1/admin/delivery-sheets?business_date=2099-01-01", headers=admin_headers)
    assert future_day.status_code == 200, future_day.text
    assert future_day.json()["date_filtered"] is True
    assert future_day.json()["units"] == []


def test_image_upload_validates_file_type(tmp_path):
    client = make_client(tmp_path)
    admin_headers, _, _, product_id = create_unit_user_product_order(client)
    bad = client.post(
        f"/api/v1/admin/products/{product_id}/image",
        headers=admin_headers,
        files={"file": ("bad.txt", b"not-image", "text/plain")},
    )
    assert bad.status_code == 400
    png = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+/p9sAAAAASUVORK5CYII="
    )
    ok = client.post(
        f"/api/v1/admin/products/{product_id}/image",
        headers=admin_headers,
        files={"file": ("image.png", png, "image/png")},
    )
    assert ok.status_code == 200
    assert ok.json()["image_path"].startswith("/uploads/")


def test_cli_resets_admin_password_without_printing_secret(tmp_path, monkeypatch, capsys):
    client = make_client(tmp_path)
    assert client.post("/api/v1/auth/login", json={"username": "root_admin", "password": "StrongPassword123"}).status_code == 200

    from app.cli import reset_admin_password

    monkeypatch.setenv("NEW_ADMIN_PASSWORD", "NewStrongPassword456")
    assert reset_admin_password("root_admin") is True

    captured = capsys.readouterr()
    output = captured.out + captured.err
    assert "NewStrongPassword456" not in output
    assert client.post("/api/v1/auth/login", json={"username": "root_admin", "password": "StrongPassword123"}).status_code == 401
    assert client.post("/api/v1/auth/login", json={"username": "root_admin", "password": "NewStrongPassword456"}).status_code == 200


def test_sessions_are_revoked_on_logout_password_reset_and_user_disable(tmp_path):
    client = make_client(tmp_path)
    admin_headers, unit_headers, _, _ = create_unit_user_product_order(client)

    assert client.get("/api/v1/auth/me", headers=unit_headers).status_code == 200
    logout = client.post("/api/v1/auth/logout", headers=unit_headers)
    assert logout.status_code == 200
    assert client.get("/api/v1/auth/me", headers=unit_headers).status_code == 401

    second_login = client.post("/api/v1/auth/login", json={"username": "unit001", "password": "UnitPassword123"})
    assert second_login.status_code == 200
    second_headers = {"Authorization": f"Bearer {second_login.json()['token']}"}
    users = client.get("/api/v1/admin/users", headers=admin_headers).json()
    unit_user = next(user for user in users if user["username"] == "unit001")
    reset = client.post(
        f"/api/v1/admin/users/{unit_user['id']}/reset-password",
        headers=admin_headers,
        json={"new_password": "ChangedPassword123", "must_change_password": True},
    )
    assert reset.status_code == 200
    assert client.get("/api/v1/auth/me", headers=second_headers).status_code == 401

    third_login = client.post("/api/v1/auth/login", json={"username": "unit001", "password": "ChangedPassword123"})
    assert third_login.status_code == 200
    third_headers = {"Authorization": f"Bearer {third_login.json()['token']}"}
    stopped = client.patch(f"/api/v1/admin/users/{unit_user['id']}/status", headers=admin_headers, json={"active": False})
    assert stopped.status_code == 200
    assert client.get("/api/v1/auth/me", headers=third_headers).status_code == 401


def test_unit_disable_blocks_login_and_order_creation_but_keeps_history(tmp_path):
    client = make_client(tmp_path)
    admin_headers, unit_headers, unit_id, product_id = create_unit_user_product_order(client)
    order = client.post(
        "/api/v1/orders",
        headers=unit_headers,
        json={"client_request_id": "REQ-unit-disable-1", "items": [{"product_id": product_id, "quantity": "1"}]},
    )
    assert order.status_code == 200

    disabled = client.patch(f"/api/v1/admin/units/{unit_id}/status", headers=admin_headers, json={"active": False})
    assert disabled.status_code == 200
    assert client.post("/api/v1/auth/login", json={"username": "unit001", "password": "UnitPassword123"}).status_code == 403
    assert client.post(
        "/api/v1/orders",
        headers=unit_headers,
        json={"client_request_id": "REQ-unit-disable-2", "items": [{"product_id": product_id, "quantity": "1"}]},
    ).status_code == 401
    history = client.get("/api/v1/admin/orders?unit_id=" + unit_id + "&include_items=true", headers=admin_headers)
    assert history.status_code == 200
    assert len(history.json()["items"]) == 1
    assert len(history.json()["items"][0]["items"]) == 1


def test_order_list_include_items_and_client_request_id_idempotency(tmp_path):
    client = make_client(tmp_path)
    admin_headers, unit_headers, _, product_id = create_unit_user_product_order(client)
    payload = {"client_request_id": "REQ-idempotent-1", "items": [{"product_id": product_id, "quantity": "2"}]}
    first = client.post("/api/v1/orders", headers=unit_headers, json=payload)
    assert first.status_code == 200, first.text
    second = client.post("/api/v1/orders", headers=unit_headers, json=payload)
    assert second.status_code == 200, second.text
    assert first.json()["id"] == second.json()["id"]
    assert first.json()["order_no"].startswith("SP")
    assert "-" in first.json()["order_no"]

    product = client.get(f"/api/v1/products/{product_id}", headers=unit_headers).json()
    assert product["reserved_quantity"] == "2"
    unit_list = client.get("/api/v1/orders?include_items=true", headers=unit_headers)
    assert unit_list.status_code == 200
    assert unit_list.json()["total"] == 1
    assert len(unit_list.json()["items"][0]["items"]) == 1
    admin_list = client.get("/api/v1/admin/orders?include_items=true&status=pending&page=1&page_size=5", headers=admin_headers)
    assert admin_list.status_code == 200
    assert admin_list.json()["total"] == 1
    assert admin_list.json()["items"][0]["client_request_id"] == "REQ-idempotent-1"


def test_migration_status_and_readiness(tmp_path):
    client = make_client(tmp_path)
    assert client.head("/api/v1/health").status_code == 200
    assert client.head("/api/v1/health/ready").status_code == 200
    ready = client.get("/api/v1/health/ready")
    assert ready.status_code == 200
    assert ready.json()["status"] == "ready"

    from app.cli import migration_status, migrate
    from app.database import connect

    status = migration_status()
    assert status["pending"] == []
    assert status["applied"]
    migrate()
    with connect() as conn:
        row = conn.execute("SELECT COUNT(*) AS c FROM schema_migrations").fetchone()
    assert row["c"] >= 1


def test_existing_database_is_migrated_without_dropping_data(tmp_path, monkeypatch):
    db_path = tmp_path / "legacy.db"
    monkeypatch.setenv("DATABASE_PATH", str(db_path))
    monkeypatch.setenv("UPLOAD_DIR", str(tmp_path / "uploads"))
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            "CREATE TABLE units(id TEXT PRIMARY KEY, unit_code TEXT UNIQUE, unit_name TEXT, default_delivery_point TEXT, active INTEGER DEFAULT 1, created_at TEXT DEFAULT CURRENT_TIMESTAMP, updated_at TEXT DEFAULT CURRENT_TIMESTAMP)"
        )
        conn.execute(
            "CREATE TABLE users(id TEXT PRIMARY KEY, username TEXT UNIQUE, password_hash TEXT, display_name TEXT, role TEXT, unit_id TEXT, active INTEGER DEFAULT 1, must_change_password INTEGER DEFAULT 0, created_at TEXT DEFAULT CURRENT_TIMESTAMP, updated_at TEXT DEFAULT CURRENT_TIMESTAMP)"
        )
        conn.execute("INSERT INTO units(id, unit_code, unit_name, default_delivery_point) VALUES ('u1', 'LEGACY', '历史单位', '历史配送点')")
        conn.commit()

    from app.database import init_db

    init_db()
    with sqlite3.connect(db_path) as conn:
        unit = conn.execute("SELECT unit_name, address_note FROM units WHERE id = 'u1'").fetchone()
        sessions = conn.execute("SELECT name FROM sqlite_master WHERE type = 'table' AND name = 'sessions'").fetchone()
    assert unit == ("历史单位", "")
    assert sessions is not None


def test_web_admin_pages_require_qr_session_and_logout_clears_cookie(tmp_path):
    client = make_client(tmp_path)
    assert client.get("/admin/dashboard", follow_redirects=False).status_code == 302
    assert client.get("/admin/dashboard", follow_redirects=False).headers["location"].startswith("/login")

    admin_headers, unit_headers, unit_id, _ = create_unit_user_product_order(client)
    admin_me = client.get("/api/v1/auth/me", headers=admin_headers).json()
    csrf = set_web_session(client, admin_me["id"], "admin")

    page = client.get("/admin/dashboard")
    assert page.status_code == 200
    assert "三公鲜配" in page.text
    assert "\u4eba\u6c11\u8b66\u5bdf\u8b66\u5fbd" not in page.text
    assert "police-badge.svg" in page.text
    assert "\u516c\u5b89" not in page.text
    assert "no-store" in page.headers["cache-control"]
    assert client.head("/admin/dashboard").status_code == 200

    admin_routes = [
        "/admin/orders",
        "/admin/products",
        "/admin/units",
        "/admin/accounts",
        "/admin/ledger",
        "/admin/receipt-issues",
        "/admin/preparation-summary",
        "/admin/delivery-sheets",
        "/admin/web-sessions",
        "/admin/system",
    ]
    for route in admin_routes:
        protected = client.get(route)
        assert protected.status_code == 200, route
        assert "三公鲜配" in protected.text
        assert "no-store" in protected.headers["cache-control"]

    me = client.get("/api/v1/web-auth/me")
    assert me.status_code == 200
    assert me.json()["role"] == "admin"

    unit_user = client.get("/api/v1/auth/me", headers=unit_headers).json()
    set_web_session(client, unit_user["id"], "unit_user", unit_id)
    forbidden = client.get("/api/v1/admin/dashboard/overview")
    assert forbidden.status_code == 403
    forbidden_page = client.get("/admin/dashboard")
    assert forbidden_page.status_code == 403
    assert "no-store" in forbidden_page.headers["cache-control"]

    csrf = set_web_session(client, admin_me["id"], "admin")
    logout = client.post("/api/v1/web-auth/logout", headers=csrf)
    assert logout.status_code == 200
    assert client.get("/api/v1/web-auth/me").status_code == 401


def test_public_help_pages_embed_role_workflow_tutorial_images(tmp_path):
    client = make_client(tmp_path)

    help_index = client.get("/help")
    admin = client.get("/help/admin")
    unit = client.get("/help/unit")

    assert help_index.status_code == 200
    assert admin.status_code == 200
    assert unit.status_code == 200
    assert "/admin-assets/workflow-admin-tutorial.png" in help_index.text
    assert "/admin-assets/workflow-unit-tutorial.png" in help_index.text
    assert "/admin-assets/workflow-admin-tutorial.png" in admin.text
    assert "/admin-assets/workflow-unit-tutorial.png" in unit.text
    assert "管理员常用流程" in admin.text
    assert "子单位常用流程" in unit.text


def test_web_qr_challenge_allows_http_public_beta_when_explicitly_enabled(tmp_path, monkeypatch):
    client = make_client(tmp_path)
    monkeypatch.setenv("APP_ENV", "production")
    monkeypatch.setenv("WEB_PUBLIC_ORIGIN", "http://47.94.227.58")
    monkeypatch.setenv("ALLOW_INSECURE_PRODUCTION_HTTP", "true")

    response = client.post("/api/v1/web-auth/qr/challenges")

    assert response.status_code == 200, response.text
    assert "secure" not in response.headers["set-cookie"].lower()
    payload = response.json()
    assert payload["challenge_id"]
    assert payload["qr_svg_data_url"].startswith("data:image/svg+xml;base64,")

    status = client.get(f"/api/v1/web-auth/qr/challenges/{payload['challenge_id']}/status")
    assert status.status_code == 200, status.text
    assert status.json()["status"] == "pending"


def test_public_beta_download_and_help_pages_are_available_without_login(tmp_path):
    client = make_client(tmp_path)

    latest = client.get("/api/v1/app-update/latest")
    assert latest.status_code == 200
    assert "available" in latest.json()

    login = client.get("/login")
    assert login.status_code == 200
    assert "下载 Android App" in login.text
    assert "三公鲜配" in login.text
    assert "undefined" not in login.text

    download = client.get("/download")
    assert download.status_code == 200
    assert "下载 Android App" in download.text
    assert "/api/v1/app-update/latest/download" in download.text
    assert "127.0.0.1" not in download.text

    for path, expected in [("/help", "首次使用"), ("/help/admin", "管理员常用流程"), ("/help/unit", "子单位常用流程")]:
        page = client.get(path)
        assert page.status_code == 200
        assert expected in page.text
        assert "景荣鲜配" not in page.text
        assert "智慧后勤采购" not in page.text


def test_admin_static_assets_avoid_cdn_storage_and_repeated_stale_label():
    admin_root = Path(__file__).resolve().parents[1] / "app" / "static" / "admin"
    unit_root = Path(__file__).resolve().parents[1] / "app" / "static" / "unit"
    dashboard_html = (admin_root / "dashboard.html").read_text(encoding="utf-8")
    dashboard_js = (admin_root / "dashboard.js").read_text(encoding="utf-8")
    login_html = (admin_root / "login.html").read_text(encoding="utf-8")
    unit_js = (unit_root / "unit.js").read_text(encoding="utf-8")

    combined_dashboard_source = dashboard_html + dashboard_js
    assert "http://" not in combined_dashboard_source
    assert "https://" not in combined_dashboard_source
    assert "localStorage" not in dashboard_js
    assert "sessionStorage" not in dashboard_js
    assert "police-badge.svg" in dashboard_html
    assert "\u4eba\u6c11\u8b66\u5bdf\u8b66\u5fbd" not in dashboard_html
    assert "\u516c\u5b89" not in dashboard_html
    assert "\u4eba\u6c11\u8b66\u5bdf\u8b66\u5fbd" not in login_html
    assert "police-badge.svg" in login_html
    assert "staleSuffix" in dashboard_js
    assert "includes(staleSuffix)" in dashboard_js
    assert 'function pageShell(title, subtitle, body = "")' in dashboard_js
    assert 'headers: { "Accept": "application/json", ...csrfHeaders, ...(options.headers || {}) },\n      ...options' not in dashboard_js
    assert 'headers: { "Accept": "application/json", ...csrfHeaders, ...(options.headers || {}) },\n      ...options' not in unit_js
    assert "expected_status" in dashboard_js
    assert "expected_version" in dashboard_js
    assert "确认发货" in dashboard_js
    assert "/ship" in dashboard_js
    assert 'name="photos"' in dashboard_js
    assert "client_request_id" in dashboard_js
    assert "下载 App" in dashboard_html
    assert "/download" in dashboard_html
    assert "/help/admin" in dashboard_html


def test_web_qr_login_is_bound_one_time_and_routes_by_server_role(tmp_path):
    client = make_client(tmp_path)
    admin_headers, unit_headers, unit_id, _ = create_unit_user_product_order(client)

    from app.database import connect
    from app.main import app
    from app.web_session import QR_BINDING_COOKIE

    with connect() as conn:
        conn.execute("UPDATE users SET must_change_password = 0 WHERE username = 'root_admin'")
        conn.commit()

    unit_challenge = client.post("/api/v1/web-auth/qr/challenges")
    assert unit_challenge.status_code == 200, unit_challenge.text
    unit_payload = unit_challenge.json()["qr_payload"]
    unit_token = unit_payload.split("token=", 1)[1]
    unit_scan = client.post(
        "/api/v1/mobile/web-auth/qr/scan",
        headers=unit_headers,
        json={"qr_token": unit_token, "device_name": "Pixel", "app_version": "1.0"},
    )
    assert unit_scan.status_code == 200, unit_scan.text
    assert unit_scan.json()["website_name"] == "三公鲜配单位网页版"
    assert unit_scan.json()["user"]["role"] == "unit_user"
    assert unit_scan.json()["user"]["unit_id"] == unit_id
    approved_unit = client.post(f"/api/v1/mobile/web-auth/qr/{unit_challenge.json()['challenge_id']}/approve", headers=unit_headers)
    assert approved_unit.status_code == 200, approved_unit.text
    consumed_unit = client.post(f"/api/v1/web-auth/qr/challenges/{unit_challenge.json()['challenge_id']}/consume")
    assert consumed_unit.status_code == 200, consumed_unit.text
    assert consumed_unit.json()["entry_url"] == "/web/entry"
    unit_entry = client.get("/web/entry", follow_redirects=False)
    assert unit_entry.status_code == 302
    assert unit_entry.headers["location"] == "/unit/home"
    assert client.get("/unit/home").status_code == 200
    assert client.get("/admin/dashboard").status_code == 403

    challenge = client.post("/api/v1/web-auth/qr/challenges", headers={"User-Agent": "Mozilla/5.0 Chrome/120 Windows"})
    assert challenge.status_code == 200, challenge.text
    challenge_data = challenge.json()
    qr_payload = challenge_data["qr_payload"]
    qr_token = qr_payload.split("token=", 1)[1]
    binding_cookie = client.cookies.get(QR_BINDING_COOKIE)
    assert binding_cookie

    with connect() as conn:
        row = conn.execute(
            "SELECT qr_token_hash, browser_binding_hash FROM web_login_challenges WHERE id = ?",
            (challenge_data["challenge_id"],),
        ).fetchone()
        assert row["qr_token_hash"] != qr_token
        assert row["browser_binding_hash"] != binding_cookie

    scan = client.post(
        "/api/v1/mobile/web-auth/qr/scan",
        headers=admin_headers,
        json={"qr_token": qr_token, "device_name": "Pixel", "app_version": "1.0"},
    )
    assert scan.status_code == 200, scan.text
    scan_data = scan.json()
    assert scan_data["challenge_id"] == challenge_data["challenge_id"]
    assert scan_data["website_name"] == "三公鲜配管理后台"
    assert scan_data["browser_name"] == "Chrome"
    assert scan_data["operating_system"] == "Windows"
    assert scan_data["ip_display"]
    assert scan_data["allowed"] is True
    assert client.get(f"/api/v1/web-auth/qr/challenges/{challenge_data['challenge_id']}/status").json()["status"] == "scanned"

    early_consume = client.post(f"/api/v1/web-auth/qr/challenges/{challenge_data['challenge_id']}/consume")
    assert early_consume.status_code == 409

    approved = client.post(f"/api/v1/mobile/web-auth/qr/{challenge_data['challenge_id']}/approve", headers=admin_headers)
    assert approved.status_code == 200, approved.text

    other_browser = TestClient(app)
    blocked = other_browser.post(f"/api/v1/web-auth/qr/challenges/{challenge_data['challenge_id']}/consume")
    assert blocked.status_code == 403

    consumed = client.post(f"/api/v1/web-auth/qr/challenges/{challenge_data['challenge_id']}/consume")
    assert consumed.status_code == 200, consumed.text
    assert consumed.json()["entry_url"] == "/web/entry"
    assert "token" not in consumed.json()
    assert "__Host-jrxp_session" not in consumed.headers.get("set-cookie", "")
    assert "httponly" in consumed.headers.get("set-cookie", "").lower()
    assert "samesite=strict" in consumed.headers.get("set-cookie", "").lower()
    admin_entry = client.get("/web/entry", follow_redirects=False)
    assert admin_entry.status_code == 302
    assert admin_entry.headers["location"] == "/admin/dashboard"

    replay = client.post(f"/api/v1/web-auth/qr/challenges/{challenge_data['challenge_id']}/consume")
    assert replay.status_code != 200


def test_unit_web_portal_cart_orders_and_isolation(tmp_path):
    client = make_client(tmp_path)
    admin_headers, unit_headers, unit_id, product_id = create_unit_user_product_order(client)
    unit_user = client.get("/api/v1/auth/me", headers=unit_headers).json()
    csrf = set_web_session(client, unit_user["id"], "unit_user", unit_id)

    assert client.get("/unit/home").status_code == 200
    assert client.get("/unit/products").status_code == 200
    assert client.get("/unit/cart").status_code == 200
    assert client.get("/unit/orders").status_code == 200
    assert client.get("/unit/profile").status_code == 200
    assert client.get("/admin/dashboard").status_code == 403
    assert client.get("/api/v1/admin/dashboard/overview").status_code == 403

    products = client.get("/unit/products/data")
    assert products.status_code == 200, products.text
    assert products.json()["items"][0]["id"] == product_id

    no_csrf = client.post("/unit/cart/items", json={"product_id": product_id, "quantity": "1"})
    assert no_csrf.status_code == 403

    added = client.post("/unit/cart/items", headers=csrf, json={"product_id": product_id, "quantity": "1"})
    assert added.status_code == 200, added.text
    cart = client.get("/unit/cart/data")
    assert cart.status_code == 200
    assert cart.json()["total_cents"] == 450
    item_id = cart.json()["items"][0]["id"]
    assert cart.json()["items"][0]["unit_id"] == unit_id

    updated = client.patch(f"/unit/cart/items/{item_id}", headers=csrf, json={"quantity": "1.5", "unit_id": "forged"})
    assert updated.status_code == 200, updated.text
    assert client.get("/unit/cart/data").json()["total_cents"] == 675

    submitted = client.post("/unit/orders", headers=csrf, json={"note": "网页下单", "unit_id": "forged"})
    assert submitted.status_code == 200, submitted.text
    order_id = submitted.json()["id"]
    assert submitted.json()["unit_id"] == unit_id
    assert submitted.json()["created_by_username"] == "unit001"
    assert client.get("/unit/cart/data").json()["items"] == []

    own_detail = client.get(f"/unit/orders/{order_id}/data")
    assert own_detail.status_code == 200
    assert own_detail.json()["id"] == order_id

    other_unit = client.post(
        "/api/v1/admin/units",
        headers=admin_headers,
        json={"unit_code": "WEB-U2", "unit_name": "其他单位", "default_delivery_point": "其他点"},
    ).json()
    other_user = client.post(
        "/api/v1/admin/users",
        headers=admin_headers,
        json={
            "username": "web_other",
            "password": "OtherPassword123",
            "display_name": "其他账号",
            "role": "unit_user",
            "unit_id": other_unit["id"],
            "must_change_password": False,
        },
    ).json()
    other_client = make_client(tmp_path)
    other_csrf = set_web_session(other_client, other_user["id"], "unit_user", other_unit["id"])
    assert other_client.get(f"/unit/orders/{order_id}/data").status_code == 404
    assert other_client.post(f"/unit/orders/{order_id}/confirm-receipt", headers=other_csrf).status_code == 404

    client.patch(f"/api/v1/admin/orders/{order_id}/status", headers=admin_headers, json={"status": "accepted"})
    client.patch(f"/api/v1/admin/orders/{order_id}/status", headers=admin_headers, json={"status": "preparing"})
    shipped = client.post(
        f"/api/v1/admin/orders/{order_id}/ship",
        headers=admin_headers,
        files=[("photos", ("proof.jpg", sample_image_bytes(), "image/jpeg"))],
        data={"note": "已发货", "client_request_id": "web-unit-ship-1"},
    )
    assert shipped.status_code == 200, shipped.text
    photo_url = shipped.json()["shipping_photos"][0]["thumbnail_url"]
    assert client.get(photo_url).status_code == 200
    assert other_client.get(photo_url).status_code == 404

    completed = client.post(f"/unit/orders/{order_id}/confirm-receipt", headers=csrf)
    assert completed.status_code == 200, completed.text
    assert completed.json()["status"] == "completed"


def test_cleanup_web_auth_records_revokes_expired_and_prunes_old_records(tmp_path):
    make_client(tmp_path)

    from app.database import connect
    from app.security import hash_token
    from app.cli import cleanup_web_auth_records

    with connect() as conn:
        user_id = conn.execute("SELECT id FROM users WHERE username = 'root_admin'").fetchone()["id"]
        conn.execute(
            """
            INSERT INTO web_login_challenges(id, qr_token_hash, browser_binding_hash, status, created_at, expires_at, consumed_at)
            VALUES ('old-challenge', 'old-token', 'old-binding', 'consumed', datetime('now', '-8 days'), datetime('now', '-8 days'), datetime('now', '-8 days'))
            """
        )
        conn.execute(
            """
            INSERT INTO web_login_challenges(id, qr_token_hash, browser_binding_hash, status, created_at, expires_at)
            VALUES ('recent-challenge', 'recent-token', 'recent-binding', 'pending', datetime('now', '-1 days'), datetime('now', '+1 minutes'))
            """
        )
        conn.execute(
            """
            INSERT INTO web_sessions(id, token_hash, user_id, role, idle_expires_at, absolute_expires_at)
            VALUES ('expired-session', ?, ?, 'admin', datetime('now', '-1 minutes'), datetime('now', '+1 hours'))
            """,
            (hash_token("expired"), user_id),
        )
        conn.execute(
            """
            INSERT INTO web_sessions(id, token_hash, user_id, role, created_at, idle_expires_at, absolute_expires_at, revoked_at, revoked_reason)
            VALUES ('old-revoked-session', ?, ?, 'admin', datetime('now', '-100 days'), datetime('now', '-99 days'), datetime('now', '-99 days'), datetime('now', '-90 days'), '测试')
            """,
            (hash_token("old-revoked"), user_id),
        )
        conn.commit()

    result = cleanup_web_auth_records(challenge_retention_days=7, revoked_session_retention_days=30)
    assert result["revoked_expired_sessions"] == 1
    assert result["deleted_challenges"] == 1
    assert result["deleted_revoked_sessions"] == 1

    with connect() as conn:
        assert conn.execute("SELECT id FROM web_login_challenges WHERE id = 'old-challenge'").fetchone() is None
        assert conn.execute("SELECT id FROM web_login_challenges WHERE id = 'recent-challenge'").fetchone() is not None
        expired = conn.execute("SELECT revoked_at, revoked_reason FROM web_sessions WHERE id = 'expired-session'").fetchone()
        assert expired["revoked_at"] is not None
        assert expired["revoked_reason"] == "会话已过期"
        assert conn.execute("SELECT id FROM web_sessions WHERE id = 'old-revoked-session'").fetchone() is None


def test_dashboard_overview_uses_shanghai_business_date_and_actual_quantity(tmp_path):
    client = make_client(tmp_path)
    admin_headers, unit_headers, unit_id, product_id = create_unit_user_product_order(client)
    admin_me = client.get("/api/v1/auth/me", headers=admin_headers).json()
    set_web_session(client, admin_me["id"], "admin")

    order_today = client.post("/api/v1/orders", headers=unit_headers, json={"items": [{"product_id": product_id, "quantity": "1"}]}).json()
    order_yesterday = client.post("/api/v1/orders", headers=unit_headers, json={"items": [{"product_id": product_id, "quantity": "1"}]}).json()
    order_cancelled = client.post("/api/v1/orders", headers=unit_headers, json={"items": [{"product_id": product_id, "quantity": "1"}]}).json()

    from app.database import connect

    with connect() as conn:
        unit_user_id = conn.execute("SELECT id FROM users WHERE username = 'unit001'").fetchone()["id"]
        conn.execute("UPDATE orders SET created_at = '2026-07-01 16:30:00', total_cents = 450 WHERE id = ?", (order_today["id"],))
        conn.execute("UPDATE orders SET created_at = '2026-07-01 15:30:00', total_cents = 450 WHERE id = ?", (order_yesterday["id"],))
        conn.execute("UPDATE orders SET created_at = '2026-07-02 01:00:00', total_cents = 450, status = 'cancelled' WHERE id = ?", (order_cancelled["id"],))
        conn.execute("UPDATE order_items SET actual_quantity = '2' WHERE order_id = ?", (order_today["id"],))
        conn.execute(
            """
            INSERT INTO receipt_issues(id, order_id, unit_id, issue_type, description, status, reported_by)
            VALUES (?, ?, ?, 'quality', '品质问题', 'open', ?)
            """,
            (str(uuid4()), order_today["id"], unit_id, unit_user_id),
        )
        conn.execute("UPDATE products SET warning_quantity = '20', supply_status = 'tight' WHERE id = ?", (product_id,))
        conn.commit()

    overview = client.get("/api/v1/admin/dashboard/overview?business_date=2026-07-02&range_days=7")
    assert overview.status_code == 200, overview.text
    body = overview.json()
    assert body["business_date"] == "2026-07-02"
    assert body["metrics"]["today_valid_orders"] == 1
    assert body["metrics"]["today_total_cents"] == 450
    assert body["metrics"]["pending"] >= 2
    assert body["metrics"]["open_receipt_issues"] == 1
    assert body["metrics"]["tight_inventory"] >= 1
    assert len(body["recent_orders"]) <= 10
    assert body["demand_rank"][0]["quantity"] == 2
    assert body["unit_rank"][0]["unit_id"] == unit_id
    assert len(body["trend"]) == 7
    assert any(item["date"] == "2026-07-02" and item["order_count"] == 1 for item in body["trend"])


def test_dashboard_overview_empty_shape_and_task_unit_labels(tmp_path):
    client = make_client(tmp_path)
    admin_headers = login(client, "root_admin", "StrongPassword123")
    admin_me = client.get("/api/v1/auth/me", headers=admin_headers).json()
    set_web_session(client, admin_me["id"], "admin")

    overview = client.get("/api/v1/admin/dashboard/overview?business_date=2026-07-02&range_days=30&unit_sort=orders")
    assert overview.status_code == 200, overview.text
    body = overview.json()

    assert body["business_date"] == "2026-07-02"
    assert len(body["trend"]) == 30
    assert body["recent_orders"] == []
    assert body["inventory_alerts"] == []
    assert body["demand_rank"] == []
    assert body["unit_rank"] == []
    assert isinstance(body["metrics"]["today_total_cents"], int)
    assert isinstance(body["metrics"]["today_valid_orders"], int)

    tasks = {item["type"]: item for item in body["tasks"]}
    assert tasks["stock_alerts"]["unit_label"] == "种"
    assert tasks["pending_orders"]["unit_label"] == "笔"
    assert tasks["waiting_shipment"]["unit_label"] == "笔"
    assert tasks["receipt_issues"]["unit_label"] == "项"


def test_dashboard_overview_inventory_metric_matches_paused_alerts(tmp_path):
    client = make_client(tmp_path)
    admin_headers = login(client, "root_admin", "StrongPassword123")
    admin_me = client.get("/api/v1/auth/me", headers=admin_headers).json()
    set_web_session(client, admin_me["id"], "admin")

    product = client.post(
        "/api/v1/admin/products",
        headers=admin_headers,
        json={
            "product_code": "PAUSED-FISH",
            "name": "暂停供应鱼丸",
            "category": "水产",
            "spec": "袋装",
            "unit": "袋",
            "price_cents": 1200,
            "stock_quantity": "100",
            "reserved_quantity": "0",
            "warning_quantity": "0",
            "supply_status": "paused",
            "active": True,
        },
    )
    assert product.status_code == 200, product.text
    product_id = product.json()["id"]

    overview = client.get("/api/v1/admin/dashboard/overview")
    assert overview.status_code == 200, overview.text
    body = overview.json()
    assert body["metrics"]["tight_inventory"] == 1
    assert body["tasks"][3]["type"] == "stock_alerts"
    assert body["tasks"][3]["count"] == 1
    assert any(item["id"] == product_id for item in body["inventory_alerts"])


def test_product_fast_entry_server_validation_matches_frontend_options(tmp_path):
    client = make_client(tmp_path)
    admin_headers = login(client, "root_admin", "StrongPassword123")
    base_payload = {
        "product_code": "FAST-VEG-001",
        "name": "快录青菜",
        "category": "蔬菜",
        "spec": "散装",
        "unit": "公斤",
        "price_cents": 320,
        "stock_quantity": "10",
        "min_order_quantity": "1",
        "quantity_step": "1",
        "warning_quantity": "0",
        "storage_method": "冷藏",
        "supply_status": "normal",
        "active": True,
    }

    ok = client.post("/api/v1/admin/products", headers=admin_headers, json=base_payload)
    assert ok.status_code == 200, ok.text

    invalid_category = client.post(
        "/api/v1/admin/products",
        headers=admin_headers,
        json={**base_payload, "product_code": "BAD-CATEGORY", "category": "办公用品"},
    )
    assert invalid_category.status_code == 400
    assert invalid_category.json()["detail"] == "食材分类不正确"

    invalid_unit = client.post(
        "/api/v1/admin/products",
        headers=admin_headers,
        json={**base_payload, "product_code": "BAD-UNIT", "unit": "随便填"},
    )
    assert invalid_unit.status_code == 400
    assert invalid_unit.json()["detail"] == "计量单位不正确"

    invalid_storage = client.post(
        "/api/v1/admin/products",
        headers=admin_headers,
        json={**base_payload, "product_code": "BAD-STORAGE", "storage_method": "露天"},
    )
    assert invalid_storage.status_code == 400
    assert invalid_storage.json()["detail"] == "储存方式不正确"

    zero_min = client.post(
        "/api/v1/admin/products",
        headers=admin_headers,
        json={**base_payload, "product_code": "BAD-MIN", "min_order_quantity": "0"},
    )
    assert zero_min.status_code == 400
    assert zero_min.json()["detail"] == "最小申领量必须大于 0"

    negative_stock = client.post(
        "/api/v1/admin/products",
        headers=admin_headers,
        json={**base_payload, "product_code": "BAD-STOCK", "stock_quantity": "-1"},
    )
    assert negative_stock.status_code == 400
    assert negative_stock.json()["detail"] == "库存不能小于 0"

    zero_price_normal = client.post(
        "/api/v1/admin/products",
        headers=admin_headers,
        json={**base_payload, "product_code": "BAD-PRICE", "price_cents": 0},
    )
    assert zero_price_normal.status_code == 400
    assert zero_price_normal.json()["detail"] == "请先填写商品价格"

    zero_price_paused = client.post(
        "/api/v1/admin/products",
        headers=admin_headers,
        json={**base_payload, "product_code": "PAUSED-ZERO", "price_cents": 0, "supply_status": "paused"},
    )
    assert zero_price_paused.status_code == 200, zero_price_paused.text

    off_shelf_status = client.patch(
        f"/api/v1/admin/products/{ok.json()['id']}/status",
        headers=admin_headers,
        json={"supply_status": "off_shelf", "active": False},
    )
    assert off_shelf_status.status_code == 400
    assert off_shelf_status.json()["detail"] == "供应状态不正确"

    available_quantity = client.post(
        "/api/v1/admin/products",
        headers=admin_headers,
        json={**base_payload, "product_code": "BAD-AVAILABLE", "available_quantity": "99"},
    )
    assert available_quantity.status_code == 422


def test_unit_invite_registration_binds_server_role_and_unit_without_storing_raw_token(tmp_path):
    client = make_client(tmp_path)
    admin_headers = login(client, "root_admin", "StrongPassword123")
    unit = client.post(
        "/api/v1/admin/units",
        headers=admin_headers,
        json={"unit_code": "REG-U1", "unit_name": "注册食堂", "default_delivery_point": "注册收货点"},
    ).json()
    other_unit = client.post(
        "/api/v1/admin/units",
        headers=admin_headers,
        json={"unit_code": "REG-U2", "unit_name": "伪造单位", "default_delivery_point": "伪造收货点"},
    ).json()

    created = client.post(
        "/api/v1/admin/invites",
        headers=admin_headers,
        json={"invite_type": "unit", "unit_id": unit["id"], "max_uses": 1, "phone_required": False},
    )
    assert created.status_code == 200, created.text
    invite_token = created.json()["invite_token"]
    assert created.json()["qr_payload"] == f"jingrongxianpei://invite?token={invite_token}"

    inspected = client.post("/api/v1/auth/invites/inspect", json={"invite_token": invite_token})
    assert inspected.status_code == 200, inspected.text
    inspect_body = inspected.json()
    assert inspect_body["valid"] is True
    assert inspect_body["invite_type"] == "unit"
    assert inspect_body["unit_name"] == "注册食堂"
    assert "role" not in inspect_body
    assert "unit_id" not in inspect_body
    assert "token_hash" not in inspect_body

    registered = client.post(
        "/api/v1/auth/register-with-invite",
        json={
            "invite_token": invite_token,
            "username": "invite_unit",
            "display_name": "邀请码子单位",
            "password": "InviteUnit123",
            "role": "admin",
            "unit_id": other_unit["id"],
            "is_admin": True,
        },
    )
    assert registered.status_code == 200, registered.text
    user = registered.json()["user"]
    assert user["role"] == "unit_user"
    assert user["unit_id"] == unit["id"]
    assert user["unit_name"] == "注册食堂"

    replay = client.post(
        "/api/v1/auth/register-with-invite",
        json={
            "invite_token": invite_token,
            "username": "invite_unit_replay",
            "display_name": "重复注册",
            "password": "InviteUnit123",
        },
    )
    assert replay.status_code == 409

    from app.database import connect
    from app.security import hash_token

    with connect() as conn:
        invite = conn.execute(
            "SELECT token_hash, display_code_suffix, used_count, status FROM registration_invites WHERE id = ?",
            (created.json()["id"],),
        ).fetchone()
        audit_count = conn.execute(
            "SELECT COUNT(*) AS c FROM audit_logs WHERE object_id = ? AND action = 'INVITE_REGISTER'",
            (created.json()["id"],),
        ).fetchone()["c"]
    assert invite["token_hash"] == hash_token(invite_token)
    assert invite["token_hash"] != invite_token
    assert invite["display_code_suffix"] == invite_token[-4:]
    assert invite["used_count"] == 1
    assert invite["status"] == "used"
    assert audit_count == 1


def test_manager_invite_registration_creates_pending_request_and_requires_approval(tmp_path):
    client = make_client(tmp_path)
    admin_headers, unit_headers, _, _ = create_unit_user_product_order(client)

    forbidden = client.post(
        "/api/v1/admin/invites",
        headers=unit_headers,
        json={"invite_type": "manager", "max_uses": 1},
    )
    assert forbidden.status_code == 403

    created = client.post(
        "/api/v1/admin/invites",
        headers=admin_headers,
        json={"invite_type": "manager", "expires_in_hours": 24, "max_uses": 1},
    )
    assert created.status_code == 200, created.text
    token = created.json()["invite_token"]
    assert created.json()["qr_payload"] == f"jingrongxianpei://invite?token={token}"

    inspected = client.post("/api/v1/auth/invites/inspect", json={"invite_token": token})
    assert inspected.status_code == 200, inspected.text
    inspect_body = inspected.json()
    assert inspect_body["valid"] is True
    assert inspect_body["invite_type"] == "manager"
    assert inspect_body["display_role"] == "管理者申请"
    assert inspect_body["approval_required"] is True
    assert "role" not in inspect_body
    assert "unit_id" not in inspect_body
    assert "token_hash" not in inspect_body
    assert "issuer_id" not in inspect_body

    registered = client.post(
        "/api/v1/auth/register-with-invite",
        json={
            "invite_token": f"jingrongxianpei://register?invite={token}",
            "username": "new_manager",
            "display_name": "新管理员",
            "password": "ManagerInvite123",
            "unit_id": "forged-unit",
        },
    )
    assert registered.status_code == 200, registered.text
    register_body = registered.json()
    assert register_body["status"] == "pending_approval"
    assert register_body["approval_required"] is True
    assert "token" not in register_body
    assert "user" not in register_body
    assert client.post("/api/v1/auth/login", json={"username": "new_manager", "password": "ManagerInvite123"}).status_code == 401

    requests = client.get("/api/v1/admin/manager-registration-requests", headers=admin_headers)
    assert requests.status_code == 200, requests.text
    pending = next(item for item in requests.json()["items"] if item["username"] == "new_manager")
    assert pending["status"] == "pending"
    assert pending["phone_masked"] == ""
    assert "password_hash" not in pending

    limited_admin = client.post(
        "/api/v1/admin/users",
        headers=admin_headers,
        json={
            "username": "limited_admin",
            "password": "LimitedAdmin123",
            "display_name": "普通管理员",
            "role": "admin",
            "must_change_password": False,
        },
    )
    assert limited_admin.status_code == 200, limited_admin.text
    from app.database import connect

    with connect() as conn:
        conn.execute("UPDATE users SET can_manage_accounts = 0 WHERE username = 'limited_admin'")
        conn.commit()
    limited_headers = login(client, "limited_admin", "LimitedAdmin123")
    blocked = client.post(f"/api/v1/admin/manager-registration-requests/{pending['id']}/approve", headers=limited_headers)
    assert blocked.status_code == 403

    approved = client.post(
        f"/api/v1/admin/manager-registration-requests/{pending['id']}/approve",
        headers=admin_headers,
        json={"review_note": "测试审批通过"},
    )
    assert approved.status_code == 200, approved.text
    assert approved.json()["status"] == "approved"
    login_after_approval = client.post("/api/v1/auth/login", json={"username": "new_manager", "password": "ManagerInvite123"})
    assert login_after_approval.status_code == 200, login_after_approval.text
    assert login_after_approval.json()["user"]["role"] == "admin"
    assert login_after_approval.json()["user"]["unit_id"] == ""


def test_system_overview_rejects_unit_users_and_returns_safety_shape(tmp_path):
    client = make_client(tmp_path)
    admin_headers, unit_headers, _, _ = create_unit_user_product_order(client)
    assert client.get("/api/v1/admin/system/overview", headers=unit_headers).status_code == 403

    overview = client.get("/api/v1/admin/system/overview", headers=admin_headers)
    assert overview.status_code == 200, overview.text
    body = overview.json()
    assert body["overall_status"] in {"healthy", "warning", "critical"}
    assert body["resources"]["scope"] in {"host", "container"}
    assert {"cpu_percent", "memory_used_bytes", "memory_total_bytes", "disk_used_bytes", "disk_total_bytes"} <= set(body["resources"])
    assert {"request_count_5m", "average_latency_ms", "p95_latency_ms", "error_count_5m", "error_rate_percent", "sqlite_lock_count_24h"} <= set(body["performance"])
    assert {"active_app_sessions", "active_web_sessions"} <= set(body["sessions"])
    assert {"database_bytes", "product_images_bytes", "shipping_photos_bytes", "receipt_issue_photos_bytes", "backups_bytes"} <= set(body["storage"])
    assert body["services"]["sms"] in {"disabled", "healthy", "unconfigured", "error"}
    assert body["capacity"]["status"] in {"sufficient", "moderate", "expand_recommended", "risk"}
    assert "压力测试" in body["capacity"]["disclaimer"]
    assert isinstance(body["alerts"], list)
    assert "APP_SECRET" not in str(body)


def test_environment_guard_exposes_only_safe_loadtest_metadata(tmp_path):
    client = make_client(tmp_path)

    production = client.get("/api/v1/system/environment")
    assert production.status_code == 200
    production_body = production.json()
    assert production_body["environment"] == "test"
    assert production_body["load_test_allowed"] is False
    assert "database_fingerprint" not in production_body
    assert "DATABASE_PATH" not in str(production_body)
    assert str(tmp_path) not in str(production_body)

    os.environ["APP_ENV"] = "loadtest"
    os.environ["LOAD_TEST_ALLOWED"] = "true"
    os.environ["DATA_NAMESPACE"] = "LOADTEST"
    loadtest = client.get("/api/v1/system/environment")
    assert loadtest.status_code == 200
    body = loadtest.json()
    assert body["environment"] == "loadtest"
    assert body["load_test_allowed"] is True
    assert body["data_namespace"] == "LOADTEST"
    assert isinstance(body["database_fingerprint"], str)
    assert len(body["database_fingerprint"]) == 16
    assert str(tmp_path) not in str(body)
