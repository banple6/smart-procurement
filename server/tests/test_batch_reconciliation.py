import json
from io import BytesIO

from openpyxl import load_workbook

from app.database import connect
from test_phase3_batches import _create_preparing_order, _create_product, _create_unit_user
from test_workflows import login, make_client


def _batch(client, headers, order_ids, name):
    response = client.post(
        "/api/v1/admin/batches",
        headers=headers,
        json={"name": name, "order_ids": order_ids, "client_request_id": f"create-{name}"},
    )
    assert response.status_code == 200, response.text
    return response.json()


def _reconcile(client, headers, unit_id, target, sources, order_ids, request_id="reconcile-one"):
    versions = {target["id"]: target["version"], **{source["id"]: source["version"] for source in sources}}
    return client.post(
        "/api/v1/admin/batches/reconcile",
        headers=headers,
        json={
            "unit_id": unit_id,
            "target_batch_id": target["id"],
            "source_batch_ids": [source["id"] for source in sources],
            "order_ids": order_ids,
            "expected_versions": versions,
            "client_request_id": request_id,
        },
    )


def test_same_unit_create_and_later_append_are_idempotent(tmp_path):
    client = make_client(tmp_path)
    admin = login(client, "root_admin", "StrongPassword123")
    product = _create_product(client, admin, "BATCH-V2-CABBAGE", "圆白菜", "斤")
    unit = _create_unit_user(client, admin, "V2-015")
    first = _create_preparing_order(client, admin, unit, product, 2)
    second = _create_preparing_order(client, admin, unit, product, 3)

    combined = _batch(client, admin, [first["id"], second["id"]], "同单位合并创建")
    repeated_create = client.post(
        "/api/v1/admin/batches",
        headers=admin,
        json={
            "name": "同单位合并创建",
            "order_ids": [first["id"], second["id"]],
            "client_request_id": "create-同单位合并创建",
        },
    )
    assert repeated_create.status_code == 200
    assert repeated_create.json()["id"] == combined["id"]

    third = _create_preparing_order(client, admin, unit, product, 4)
    workbench = client.get("/api/v1/admin/batches/unit-workbench", headers=admin).json()["items"]
    group = next(item for item in workbench if item["unit_id"] == first["unit_id"])
    assert group["pending_order_count"] == 1
    assert group["open_batch_count"] == 1
    assert group["open_batches"][0]["id"] == combined["id"]

    payload = {
        "add_order_ids": [third["id"]],
        "remove_order_ids": [],
        "expected_version": combined["version"],
        "client_request_id": "append-third",
    }
    appended = client.patch(f"/api/v1/admin/batches/{combined['id']}/orders", headers=admin, json=payload)
    assert appended.status_code == 200, appended.text
    assert {order["id"] for order in appended.json()["orders"]} == {first["id"], second["id"], third["id"]}
    repeated = client.patch(f"/api/v1/admin/batches/{combined['id']}/orders", headers=admin, json=payload)
    assert repeated.status_code == 200
    assert len(repeated.json()["orders"]) == 3

    stale = client.patch(
        f"/api/v1/admin/batches/{combined['id']}/orders",
        headers=admin,
        json={
            "remove_order_ids": [third["id"]],
            "expected_version": combined["version"],
            "client_request_id": "stale-remove",
        },
    )
    assert stale.status_code == 409
    with connect() as conn:
        assert conn.execute(
            "SELECT COUNT(*) FROM delivery_batch_orders WHERE order_id = ?", (third["id"],)
        ).fetchone()[0] == 1
        audit_rows = conn.execute(
            "SELECT action, before_json, after_json FROM audit_logs WHERE object_id = ?", (combined["id"],)
        ).fetchall()
        actions = {row[0] for row in audit_rows}
        update_audit = next(row for row in audit_rows if row[0] == "DELIVERY_BATCH_ORDERS_UPDATED")
    assert {"DELIVERY_BATCH_ORDERS_UPDATED", "BATCH_ORDER_ADDED"} <= actions
    assert set(json.loads(update_audit[1])["order_ids"]) == {first["id"], second["id"]}
    assert set(json.loads(update_audit[2])["order_ids"]) == {first["id"], second["id"], third["id"]}


def test_reconcile_moves_only_requested_unit_and_preserves_mixed_source(tmp_path):
    client = make_client(tmp_path)
    admin = login(client, "root_admin", "StrongPassword123")
    jin = _create_product(client, admin, "BATCH-V2-JIN", "鸡蛋", "斤")
    box = _create_product(client, admin, "BATCH-V2-BOX", "鸡蛋", "盒")
    unit_a = _create_unit_user(client, admin, "V2-A")
    unit_b = _create_unit_user(client, admin, "V2-B")
    order_a1 = _create_preparing_order(client, admin, unit_a, jin, 20)
    order_a2 = _create_preparing_order(client, admin, unit_a, box, 5)
    order_b = _create_preparing_order(client, admin, unit_b, jin, 7)
    target = _batch(client, admin, [order_a1["id"]], "目标单")
    source = _batch(client, admin, [order_a2["id"], order_b["id"]], "混合来源单")

    before = {}
    with connect() as conn:
        for order_id in (order_a1["id"], order_a2["id"], order_b["id"]):
            before[order_id] = tuple(
                conn.execute(
                    """
                    SELECT product_name_snapshot, spec_snapshot, unit_snapshot, quantity,
                           price_cents_snapshot, subtotal_cents
                    FROM order_items WHERE order_id = ? ORDER BY rowid
                    """,
                    (order_id,),
                ).fetchone()
            )

    wrong_unit = _reconcile(
        client,
        admin,
        order_a1["unit_id"],
        target,
        [source],
        [order_a1["id"], order_b["id"]],
        "cross-unit-rejected",
    )
    assert wrong_unit.status_code == 409

    merged = _reconcile(
        client,
        admin,
        order_a1["unit_id"],
        target,
        [source],
        [order_a1["id"], order_a2["id"]],
    )
    assert merged.status_code == 200, merged.text
    assert {order["id"] for order in merged.json()["orders"]} == {order_a1["id"], order_a2["id"]}
    source_after = client.get(f"/api/v1/admin/batches/{source['id']}", headers=admin).json()
    assert source_after["status"] == "open"
    assert [order["id"] for order in source_after["orders"]] == [order_b["id"]]

    summary = client.get(f"/api/v1/admin/batches/{target['id']}/summary", headers=admin).json()
    egg_lines = [item for item in summary["by_product"] if item["product_name"] == "鸡蛋"]
    assert {(line["total_quantity"], line["unit"]) for line in egg_lines} == {("20", "斤"), ("5", "盒")}
    exported = client.get(f"/api/v1/admin/batches/{target['id']}/picking-list.xlsx", headers=admin)
    assert exported.status_code == 200
    workbook = load_workbook(BytesIO(exported.content), data_only=True, read_only=True)
    values = [cell for row in workbook[workbook.sheetnames[0]].iter_rows(values_only=True) for cell in row]
    assert "鸡蛋" in values

    with connect() as conn:
        after = {}
        for order_id in before:
            after[order_id] = tuple(
                conn.execute(
                    """
                    SELECT product_name_snapshot, spec_snapshot, unit_snapshot, quantity,
                           price_cents_snapshot, subtotal_cents
                    FROM order_items WHERE order_id = ? ORDER BY rowid
                    """,
                    (order_id,),
                ).fetchone()
            )
        audit = conn.execute(
            "SELECT before_json, after_json FROM audit_logs WHERE action = 'BATCH_RECONCILED' AND object_id = ?",
            (target["id"],),
        ).fetchone()
    assert after == before
    assert audit is not None
    assert json.loads(audit[1])["moved_order_ids"] == [order_a2["id"]]


def test_reconcile_archives_empty_source_and_is_idempotent(tmp_path):
    client = make_client(tmp_path)
    admin = login(client, "root_admin", "StrongPassword123")
    product = _create_product(client, admin, "BATCH-V2-EMPTY", "土豆", "斤")
    unit = _create_unit_user(client, admin, "V2-EMPTY")
    first = _create_preparing_order(client, admin, unit, product, 2)
    second = _create_preparing_order(client, admin, unit, product, 3)
    target = _batch(client, admin, [first["id"]], "保留目标")
    source = _batch(client, admin, [second["id"]], "归档来源")

    response = _reconcile(
        client,
        admin,
        first["unit_id"],
        target,
        [source],
        [first["id"], second["id"]],
        "archive-empty-source",
    )
    assert response.status_code == 200, response.text
    repeated = _reconcile(
        client,
        admin,
        first["unit_id"],
        target,
        [source],
        [first["id"], second["id"]],
        "archive-empty-source",
    )
    assert repeated.status_code == 200, repeated.text
    source_after = client.get(f"/api/v1/admin/batches/{source['id']}", headers=admin).json()
    assert source_after["status"] == "cancelled"
    assert source_after["orders"] == []
    with connect() as conn:
        assert conn.execute(
            "SELECT COUNT(*) FROM audit_logs WHERE action = 'BATCH_RECONCILED' AND request_id = ?",
            ("archive-empty-source",),
        ).fetchone()[0] == 1
        assert conn.execute(
            "SELECT COUNT(*) FROM delivery_batch_orders WHERE order_id = ?", (second["id"],)
        ).fetchone()[0] == 1


def test_adjustment_rejects_closed_and_outbound_batches_and_remove_returns_pool(tmp_path):
    client = make_client(tmp_path)
    admin = login(client, "root_admin", "StrongPassword123")
    product = _create_product(client, admin, "BATCH-V2-LOCK", "茄子", "斤")
    unit = _create_unit_user(client, admin, "V2-LOCK")
    first = _create_preparing_order(client, admin, unit, product, 2)
    second = _create_preparing_order(client, admin, unit, product, 3)
    batch = _batch(client, admin, [first["id"], second["id"]], "调整测试")

    removed = client.patch(
        f"/api/v1/admin/batches/{batch['id']}/orders",
        headers=admin,
        json={
            "remove_order_ids": [second["id"]],
            "expected_version": batch["version"],
            "client_request_id": "remove-second",
        },
    )
    assert removed.status_code == 200, removed.text
    eligible_ids = {row["id"] for row in client.get("/api/v1/admin/batches/eligible-orders", headers=admin).json()["items"]}
    assert second["id"] in eligible_ids

    closed = client.patch(
        f"/api/v1/admin/batches/{batch['id']}/status",
        headers=admin,
        json={"status": "closed", "expected_version": removed.json()["version"]},
    )
    assert closed.status_code == 200
    blocked_closed = client.patch(
        f"/api/v1/admin/batches/{batch['id']}/orders",
        headers=admin,
        json={"add_order_ids": [second["id"]], "expected_version": closed.json()["version"], "client_request_id": "closed-add"},
    )
    assert blocked_closed.status_code == 409

    outbound = client.post(f"/api/v1/admin/outbounds/from-batch/{batch['id']}", headers=admin)
    assert outbound.status_code == 200, outbound.text
    with connect() as conn:
        conn.execute("UPDATE delivery_batches SET status = 'open' WHERE id = ?", (batch["id"],))
        conn.commit()
    blocked_outbound = client.patch(
        f"/api/v1/admin/batches/{batch['id']}/orders",
        headers=admin,
        json={"add_order_ids": [second["id"]], "expected_version": closed.json()["version"], "client_request_id": "outbound-add"},
    )
    assert blocked_outbound.status_code == 409
    assert "出库单" in blocked_outbound.json()["detail"]
