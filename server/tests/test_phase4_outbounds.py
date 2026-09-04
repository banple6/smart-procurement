from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
from zipfile import ZipFile

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from PIL import Image

from test_phase3_batches import _create_preparing_order, _create_product, _create_unit_user
from test_workflows import login, make_client
from app.services.batch_exports import outbound_order_workbook


def _png_bytes() -> bytes:
    stream = BytesIO()
    Image.new("RGB", (64, 48), "green").save(stream, format="PNG")
    return stream.getvalue()


def _setup_closed_batch(client, admin_headers):
    product_jin = _create_product(client, admin_headers, "PH4-POTATO-JIN", "Phase4土豆", "斤")
    product_box = _create_product(client, admin_headers, "PH4-POTATO-BOX", "Phase4土豆", "箱")
    unit_a = _create_unit_user(client, admin_headers, "PH4A")
    unit_b = _create_unit_user(client, admin_headers, "PH4B")
    unit_c = _create_unit_user(client, admin_headers, "PH4C")
    orders = [
        _create_preparing_order(client, admin_headers, unit_a, product_jin, 3),
        _create_preparing_order(client, admin_headers, unit_a, product_jin, 2),
        _create_preparing_order(client, admin_headers, unit_b, product_jin, 7),
        _create_preparing_order(client, admin_headers, unit_b, product_box, 2),
        _create_preparing_order(client, admin_headers, unit_c, product_jin, 4),
    ]
    batch = client.post(
        "/api/v1/admin/batches",
        headers=admin_headers,
        json={"name": "Phase4单位出库", "note": "隔离测试", "order_ids": [order["id"] for order in orders]},
    )
    assert batch.status_code == 200, batch.text
    closed = client.patch(
        f"/api/v1/admin/batches/{batch.json()['id']}/status",
        headers=admin_headers,
        json={"status": "closed", "expected_version": batch.json()["version"]},
    )
    assert closed.status_code == 200, closed.text
    return batch.json(), orders, (unit_a, unit_b, unit_c), (product_jin, product_box)


def _generate(client, admin_headers, batch_id):
    response = client.post(f"/api/v1/admin/outbounds/from-batch/{batch_id}", headers=admin_headers)
    assert response.status_code == 200, response.text
    return response.json()


def _worksheet_values(sheet):
    return [str(cell.value or "") for row in sheet.iter_rows() for cell in row]


def _assert_final_outbound_layout(sheet):
    values = _worksheet_values(sheet)
    joined = "\n".join(values)
    assert any(value.startswith("三公鲜配出库单（") for value in values)
    assert any(value.startswith("单位：") for value in values)
    assert any(value.startswith("系统出库单号：") for value in values)
    assert any(value.startswith("日期：") for value in values)
    assert values.count("总金额：") == 1
    assert any(value.startswith("配送人：") for value in values)
    assert any(value.startswith("收货人：") for value in values)
    for removed in ("来源备货单", "出库人", "分类金额汇总", "日期：________________"):
        assert removed not in joined
    assert sheet.page_setup.fitToWidth == 1


def test_outbound_workbook_final_layout_uses_snapshot_cents_and_keeps_metadata_readable():
    workbook = load_workbook(
        BytesIO(
            outbound_order_workbook(
                    {
                        "unit_code": "006",
                        "unit_name_snapshot": "新集派出所",
                        "outbound_no": "CK20260903-0018",
                        "created_at": "2026-09-03 16:56:00",
                        "business_date": "2026-09-03",
                        "batch_no": "PS20260903-0001",
                },
                [
                    {"category": "水果", "product_name": "苹果", "spec": "散装", "unit": "公斤", "quantity": "5", "price_cents_snapshot": 2000, "subtotal_cents": 10000},
                    {"category": "蔬菜", "product_name": "西红柿", "spec": "散装", "unit": "公斤", "quantity": "4", "price_cents_snapshot": 1500, "subtotal_cents": 6000},
                ],
            )
        ),
        data_only=True,
    )
    sheet = workbook["出库单"]

    assert sheet["A1"].value == "三公鲜配出库单（20260903/新集派出所）"
    assert sheet["A2"].value == "单位：006 · 新集派出所"
    assert sheet["D2"].value == "系统出库单号：CK20260903-0018"
    assert sheet["G2"].value == "日期：2026-09-03 16:56:00"
    assert {str(cell_range) for cell_range in sheet.merged_cells.ranges} >= {"A2:C2", "D2:F2", "G2:H2"}
    assert sheet.column_dimensions["G"].width + sheet.column_dimensions["H"].width >= 30
    assert [sheet.cell(3, column).value for column in range(1, 9)] == ["序号", "食品分类", "食材名称", "规格", "计量单位", "需求数量", "单价", "小计金额"]
    assert [sheet.cell(4, column).value for column in range(1, 9)] == [1, "水果", "苹果", "散装", "公斤", "5", 20, 100]
    assert [sheet.cell(5, column).value for column in range(1, 9)] == [2, "蔬菜", "西红柿", "散装", "公斤", "4", 15, 60]
    assert sheet["G4"].data_type == "n"
    assert sheet["H4"].data_type == "n"
    assert sheet["G6"].value == "总金额："
    assert sheet["H6"].value == 160
    assert sheet["G4"].number_format == '¥0.00'
    assert sheet["H4"].number_format == '¥0.00'
    assert sheet["H6"].number_format == '¥0.00'
    assert sheet["A8"].value == "配送人：________________"
    assert sheet["E8"].value == "收货人：________________"
    assert str(sheet.print_area).endswith("!$A$1:$H$8")
    _assert_final_outbound_layout(sheet)

    precision_workbook = load_workbook(
        BytesIO(
            outbound_order_workbook(
                {"unit_name_snapshot": "测试单位02", "outbound_no": "CK20260827-0002", "created_at": "2026-08-27 16:56:00"},
                [
                    {"category": "蔬菜", "product_name": "A", "spec": "", "unit": "公斤", "quantity": "1.5", "price_cents_snapshot": 214, "subtotal_cents": 321},
                    {"category": "蔬菜", "product_name": "B", "spec": "", "unit": "公斤", "quantity": "1", "price_cents_snapshot": 107, "subtotal_cents": 107},
                    {"category": "其他", "product_name": "C", "spec": "", "unit": "公斤", "quantity": "1", "price_cents_snapshot": 999, "subtotal_cents": 999},
                ],
            )
        ),
        data_only=True,
    )
    precision_sheet = precision_workbook["出库单"]
    assert precision_sheet["G4"].value == 2.14
    assert precision_sheet["H7"].value == 14.27
    assert precision_sheet["H7"].number_format == '¥0.00'

    zero_workbook = load_workbook(
        BytesIO(
            outbound_order_workbook(
                {"unit_name_snapshot": "测试单位03", "outbound_no": "CK20260827-0003", "created_at": "2026-08-27 16:56:00"},
                [],
            )
        ),
        data_only=True,
    )
    assert zero_workbook["出库单"]["H4"].value == 0
    assert zero_workbook["出库单"]["H4"].number_format == '¥0.00'

    long_name_workbook = load_workbook(
        BytesIO(
            outbound_order_workbook(
                {
                    "unit_code": "002",
                    "unit_name_snapshot": "三河市公安局燕郊分局",
                    "outbound_no": "CK20260903-0019",
                    "business_date": "2026-09-03",
                },
                [],
            )
        ),
        data_only=True,
    )
    assert long_name_workbook["出库单"]["A1"].value == "三公鲜配出库单（20260903/三河市公安局燕郊分局）"

    fallback_workbook = load_workbook(
        BytesIO(
            outbound_order_workbook(
                {"unit_code": "006", "outbound_no": "CK20260903-0020", "business_date": "2026-09-03"},
                [],
            )
        ),
        data_only=True,
    )
    assert fallback_workbook["出库单"]["A1"].value == "三公鲜配出库单（20260903/006）"
    assert fallback_workbook["出库单"]["A2"].value == "单位：006 · 未命名单位"


def test_outbound_workbook_amount_columns_keep_historical_prices_and_blank_missing_prices():
    workbook = load_workbook(
        BytesIO(
            outbound_order_workbook(
                {
                    "unit_code": "016",
                    "unit_name_snapshot": "高楼派出所",
                    "outbound_no": "CK20260904-0006",
                    "business_date": "2026-09-04",
                    "batch_no": "PS20260904-0003",
                },
                [
                    {"category": "调料", "product_name": "八角", "spec": "散装", "unit": "斤", "quantity": "2", "price_cents_snapshot": 3159, "subtotal_cents": 6318, "current_product_price_cents": 9999},
                    {"category": "调料", "product_name": "孜然(粒)", "spec": "预包装500G", "unit": "袋", "quantity": "1", "price_cents_snapshot": 2106, "subtotal_cents": 2106},
                    {"category": "调料", "product_name": "特级云南花椒", "spec": "原装特级长规格包装", "unit": "袋", "quantity": "1", "price_cents_snapshot": 18396, "subtotal_cents": 18396},
                ],
            )
        ),
        data_only=True,
    )
    sheet = workbook["出库单"]

    assert sheet["A1"].value == "三公鲜配出库单（20260904/高楼派出所）"
    assert sheet["A2"].value == "单位：016 · 高楼派出所"
    assert sheet["D2"].value == "系统出库单号：CK20260904-0006"
    assert [sheet.cell(3, column).value for column in range(1, 9)] == ["序号", "食品分类", "食材名称", "规格", "计量单位", "需求数量", "单价", "小计金额"]
    assert [sheet.cell(4, column).value for column in range(1, 9)] == [1, "调料", "八角", "散装", "斤", "2", 31.59, 63.18]
    assert [sheet.cell(5, column).value for column in range(1, 9)] == [2, "调料", "孜然(粒)", "预包装500G", "袋", "1", 21.06, 21.06]
    assert sheet["G4"].data_type == "n"
    assert sheet["H4"].data_type == "n"
    assert sheet["G4"].number_format == '¥0.00'
    assert sheet["H4"].number_format == '¥0.00'
    assert sheet["G7"].value == "总金额："
    assert sheet["H7"].value == 268.2
    assert sum(sheet.cell(row, 8).value for row in range(4, 7)) == sheet["H7"].value
    assert str(sheet.print_area).endswith("!$A$1:$H$9")
    assert sheet["A9"].value == "配送人：________________"
    assert sheet["E9"].value == "收货人：________________"

    missing_price_workbook = load_workbook(
        BytesIO(
            outbound_order_workbook(
                {"unit_code": "016", "unit_name_snapshot": "高楼派出所", "outbound_no": "CK20260904-0007", "business_date": "2026-09-04"},
                [{"category": "调料", "product_name": "历史缺价食材", "spec": "散装", "unit": "斤", "quantity": "2", "subtotal_cents": 6318}],
            )
        ),
        data_only=True,
    )
    missing_price_sheet = missing_price_workbook["出库单"]
    assert missing_price_sheet["G4"].value is None
    assert missing_price_sheet["H4"].value == 63.18


def test_phase4_outbounds_split_units_export_and_keep_snapshot(tmp_path):
    client = make_client(tmp_path)
    admin_headers = login(client, "root_admin", "StrongPassword123")
    batch, orders, _, products = _setup_closed_batch(client, admin_headers)

    generated = _generate(client, admin_headers, batch["id"])
    assert generated["created_count"] == 3
    assert len(generated["items"]) == 3
    by_name = {item["unit_name_snapshot"]: item for item in generated["items"]}
    unit_a = next(item for item in generated["items"] if item["order_count"] == 2)
    detail = client.get(f"/api/v1/admin/outbounds/{unit_a['id']}", headers=admin_headers)
    assert detail.status_code == 200, detail.text
    body = detail.json()
    assert body["total_cents"] == sum(line["subtotal_cents"] for line in body["lines"])
    assert len(body["orders"]) == 2
    assert {order["id"] for order in body["orders"]} == {orders[0]["id"], orders[1]["id"]}
    assert [(line["quantity"], line["unit"]) for line in body["lines"]] == [("5", "斤")]

    listed = client.get("/api/v1/admin/outbounds", headers=admin_headers)
    assert listed.status_code == 200, listed.text
    listed_item = next(item for item in listed.json()["items"] if item["id"] == unit_a["id"])
    assert listed_item["total_cents"] == body["total_cents"]

    summary = client.get(f"/api/v1/admin/batches/{batch['id']}/summary", headers=admin_headers).json()
    assert sum(float(line["quantity"]) for item in generated["items"] for line in client.get(f"/api/v1/admin/outbounds/{item['id']}", headers=admin_headers).json()["lines"] if line["unit"] == "斤") == sum(float(item["total_quantity"]) for item in summary["by_product"] if item["unit"] == "斤")

    current = client.get(f"/api/v1/products/{products[0]}", headers=admin_headers).json()
    changed = client.patch(f"/api/v1/admin/products/{products[0]}/price", headers=admin_headers, json={"price_cents": 999, "expected_version": current["version"]})
    assert changed.status_code == 200, changed.text
    exported = client.get(f"/api/v1/admin/outbounds/{unit_a['id']}/export.xlsx", headers=admin_headers)
    assert exported.status_code == 200, exported.text
    assert len(exported.content) > 0
    workbook = load_workbook(BytesIO(exported.content), data_only=True)
    sheet = workbook["出库单"]
    assert sheet.cell(1, 1).value.startswith(f"三公鲜配出库单（")
    assert sheet.cell(1, 1).value.endswith(f"/{unit_a['unit_name_snapshot']}）")
    assert sheet["A2"].value.startswith(f"单位：{body['unit_code']} · ")
    assert unit_a["outbound_no"] in " ".join(str(cell.value or "") for cell in sheet[2])
    assert sheet["F4"].value == "5"
    assert sheet["F4"].number_format == "0.###"
    assert sheet["G4"].value == 1
    assert sheet["G4"].value != 9.99
    assert sheet["H4"].value == 5
    assert sheet["H4"].number_format == '¥0.00'
    _assert_final_outbound_layout(sheet)

    bulk = client.get("/api/v1/admin/outbounds/bulk.zip?" + "&".join(f"outbound_ids={item['id']}" for item in generated["items"]), headers=admin_headers)
    assert bulk.status_code == 200, bulk.text
    with ZipFile(BytesIO(bulk.content)) as archive:
        assert len(archive.namelist()) == 3
        for filename in archive.namelist():
            assert filename.endswith(".xlsx")
            workbook = load_workbook(BytesIO(archive.read(filename)), data_only=True)
            assert workbook.sheetnames == ["出库单"]
            _assert_final_outbound_layout(workbook["出库单"])

    repeated = _generate(client, admin_headers, batch["id"])
    assert repeated["created_count"] == 0
    assert {item["id"] for item in repeated["items"]} == {item["id"] for item in generated["items"]}
    assert by_name


def test_outbound_export_keeps_same_product_lines_separate_by_historical_unit_price(tmp_path):
    client = make_client(tmp_path)
    admin_headers = login(client, "root_admin", "StrongPassword123")
    product_id = _create_product(client, admin_headers, "PH4-HISTORICAL-PRICE", "历史单价食材", "斤")
    unit_headers = _create_unit_user(client, admin_headers, "HISTORICAL-PRICE")
    first = _create_preparing_order(client, admin_headers, unit_headers, product_id, 2)

    current = client.get(f"/api/v1/products/{product_id}", headers=admin_headers).json()
    changed = client.patch(
        f"/api/v1/admin/products/{product_id}/price",
        headers=admin_headers,
        json={"price_cents": 3159, "expected_version": current["version"]},
    )
    assert changed.status_code == 200, changed.text
    second = _create_preparing_order(client, admin_headers, unit_headers, product_id, 1)
    batch = client.post(
        "/api/v1/admin/batches",
        headers=admin_headers,
        json={"name": "历史单价导出", "order_ids": [first["id"], second["id"]]},
    )
    assert batch.status_code == 200, batch.text
    closed = client.patch(
        f"/api/v1/admin/batches/{batch.json()['id']}/status",
        headers=admin_headers,
        json={"status": "closed", "expected_version": batch.json()["version"]},
    )
    assert closed.status_code == 200, closed.text
    outbound = _generate(client, admin_headers, batch.json()["id"])["items"][0]

    detail = client.get(f"/api/v1/admin/outbounds/{outbound['id']}", headers=admin_headers)
    assert detail.status_code == 200, detail.text
    lines = detail.json()["lines"]
    assert [(line["quantity"], line["price_cents_snapshot"], line["subtotal_cents"]) for line in lines] == [
        ("2", 100, 200),
        ("1", 3159, 3159),
    ]
    assert detail.json()["total_cents"] == 3359

    exported = client.get(f"/api/v1/admin/outbounds/{outbound['id']}/export.xlsx", headers=admin_headers)
    assert exported.status_code == 200, exported.text
    sheet = load_workbook(BytesIO(exported.content), data_only=True)["出库单"]
    assert [(sheet.cell(row, 6).value, sheet.cell(row, 7).value, sheet.cell(row, 8).value) for row in (4, 5)] == [
        ("2", 1, 2),
        ("1", 31.59, 31.59),
    ]
    assert sheet["H6"].value == 33.59


def test_phase4_rejects_open_batch_and_is_concurrent_safe(tmp_path):
    client = make_client(tmp_path)
    admin_headers = login(client, "root_admin", "StrongPassword123")
    product_id = _create_product(client, admin_headers, "PH4-OPEN", "Phase4白菜", "斤")
    unit = _create_unit_user(client, admin_headers, "PH4OPEN")
    order = _create_preparing_order(client, admin_headers, unit, product_id, 2)
    batch = client.post("/api/v1/admin/batches", headers=admin_headers, json={"name": "未完成备货", "order_ids": [order["id"]]}).json()
    assert client.post(f"/api/v1/admin/outbounds/from-batch/{batch['id']}", headers=admin_headers).status_code == 409
    client.patch(f"/api/v1/admin/batches/{batch['id']}/status", headers=admin_headers, json={"status": "closed", "expected_version": batch["version"]})

    from app.main import app

    client_b = TestClient(app)
    headers_b = login(client_b, "root_admin", "StrongPassword123")
    with ThreadPoolExecutor(max_workers=2) as executor:
        results = list(executor.map(lambda args: args[0].post(f"/api/v1/admin/outbounds/from-batch/{batch['id']}", headers=args[1]).status_code, [(client, admin_headers), (client_b, headers_b)]))
    assert sorted(results) == [200, 200]
    listed = client.get("/api/v1/admin/outbounds", headers=admin_headers).json()["items"]
    assert len(listed) == 1


def test_phase4_ships_only_selected_unit_once_and_preserves_receipt_flow(tmp_path):
    client = make_client(tmp_path)
    admin_headers = login(client, "root_admin", "StrongPassword123")
    batch, orders, unit_headers, products = _setup_closed_batch(client, admin_headers)
    generated = _generate(client, admin_headers, batch["id"])
    unit_a = next(item for item in generated["items"] if item["order_count"] == 2)
    before = client.get(f"/api/v1/products/{products[0]}", headers=admin_headers).json()
    payload = {
        "data": {"note": "Phase4 发货核对", "client_request_id": "phase4-ship-a"},
        "files": [("photos", ("proof.png", _png_bytes(), "image/png"))],
    }
    shipped = client.post(f"/api/v1/admin/outbounds/{unit_a['id']}/ship", headers=admin_headers, **payload)
    assert shipped.status_code == 200, shipped.text
    assert shipped.json()["status"] == "shipped"
    statuses = {order["id"]: client.get(f"/api/v1/admin/orders/{order['id']}", headers=admin_headers).json()["status"] for order in orders}
    assert statuses[orders[0]["id"]] == "shipped"
    assert statuses[orders[1]["id"]] == "shipped"
    assert statuses[orders[2]["id"]] == "preparing"
    assert statuses[orders[3]["id"]] == "preparing"
    assert statuses[orders[4]["id"]] == "preparing"
    assert all(client.get(f"/api/v1/admin/orders/{order['id']}", headers=admin_headers).json()["shipping_photo_count"] == 1 for order in orders[:2])
    after_ship = client.get(f"/api/v1/products/{products[0]}", headers=admin_headers).json()
    assert (after_ship["stock_quantity"], after_ship["reserved_quantity"]) == (before["stock_quantity"], before["reserved_quantity"])
    retry = client.post(f"/api/v1/admin/outbounds/{unit_a['id']}/ship", headers=admin_headers, **payload)
    assert retry.status_code == 200, retry.text
    assert client.get(f"/api/v1/admin/orders/{orders[0]['id']}", headers=admin_headers).json()["shipping_photo_count"] == 1

    receipt = client.post(f"/api/v1/orders/{orders[0]['id']}/confirm-receipt", headers=unit_headers[0])
    assert receipt.status_code == 200, receipt.text
    assert receipt.json()["status"] == "completed"
    after_receipt = client.get(f"/api/v1/products/{products[0]}", headers=admin_headers).json()
    assert after_receipt["stock_quantity"] == "997"
    assert after_receipt["reserved_quantity"] == "13"

    assert client.delete(f"/api/v1/admin/outbounds/{unit_a['id']}", headers=admin_headers).status_code == 409
    pending = next(item for item in generated["items"] if item["id"] != unit_a["id"])
    archived = client.delete(f"/api/v1/admin/outbounds/{pending['id']}", headers=admin_headers)
    assert archived.status_code == 200, archived.text
    assert archived.json()["status"] == "archived"
