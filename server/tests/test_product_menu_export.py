from io import BytesIO
from numbers import Number
from urllib.parse import unquote

from openpyxl import load_workbook

from test_new_business_features import create_product, create_unit_order, login, make_client


def test_admin_product_menu_export_uses_active_catalog_and_current_prices(tmp_path):
    client = make_client(tmp_path)
    admin_headers = login(client, "root_admin", "StrongPassword123")
    tomato = create_product(client, admin_headers, code="MENU-TOMATO", name="西红柿", unit="斤", price_cents=199)
    egg = create_product(client, admin_headers, code="MENU-EGG", name="鸡蛋", unit="斤", price_cents=580)
    paused = create_product(client, admin_headers, code="MENU-PAUSED", name="已停用食材", unit="斤", price_cents=320)
    archived = create_product(client, admin_headers, code="MENU-ARCHIVED", name="已归档食材", unit="斤", price_cents=450)

    order = create_unit_order(client, admin_headers, tomato["id"], "MENU", "1")
    tomato_after_order = client.get(f"/api/v1/products/{tomato['id']}", headers=admin_headers)
    assert tomato_after_order.status_code == 200, tomato_after_order.text
    current_tomato = client.patch(
        f"/api/v1/admin/products/{tomato['id']}/price",
        headers=admin_headers,
        json={"price_cents": 250, "expected_version": tomato_after_order.json()["version"]},
    )
    assert current_tomato.status_code == 200, current_tomato.text
    paused_response = client.patch(
        f"/api/v1/admin/products/{paused['id']}/status",
        headers=admin_headers,
        json={"supply_status": "paused", "active": False, "expected_version": paused["version"]},
    )
    assert paused_response.status_code == 200, paused_response.text
    archived_response = client.delete(f"/api/v1/admin/products/{archived['id']}", headers=admin_headers)
    assert archived_response.status_code == 200, archived_response.text
    unit_headers = login(client, "unit_MENU", "UnitMENUPassword123")

    from app.database import connect

    with connect() as conn:
        before = {
            table: conn.execute(f"SELECT COUNT(*) AS c FROM {table}").fetchone()["c"]
            for table in ("products", "orders", "order_items", "audit_logs")
        }
        snapshot_price = conn.execute(
            "SELECT price_cents_snapshot FROM order_items WHERE order_id = ?", (order["id"],)
        ).fetchone()["price_cents_snapshot"]
    assert snapshot_price == 199

    exported = client.get("/api/v1/admin/products/export.xlsx", headers=admin_headers)
    assert exported.status_code == 200, exported.text
    assert exported.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "filename*=UTF-8''" in exported.headers["content-disposition"]
    assert "三公鲜配商品菜单_" in unquote(exported.headers["content-disposition"])

    workbook = load_workbook(BytesIO(exported.content), data_only=False)
    assert workbook.properties.title == "三公鲜配商品菜单"
    assert workbook.sheetnames == ["三公鲜配商品菜单"]
    sheet = workbook.active
    assert [cell.value for cell in sheet[1]] == ["序号", "商品名称", "规格", "当前价格"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    expected_menu = client.get("/api/v1/products", headers=unit_headers)
    assert expected_menu.status_code == 200, expected_menu.text
    assert [(row[1], row[2], row[3]) for row in rows] == [
        (item["name"], item["spec"], item["price_cents"] / 100) for item in expected_menu.json()
    ]
    assert {(row[1], row[2], row[3]) for row in rows} == {("西红柿", "散装", 2.5), ("鸡蛋", "散装", 5.8)}
    assert all(isinstance(row[3], Number) for row in rows)
    assert all(row[1] not in {"已停用食材", "已归档食材"} for row in rows)

    with connect() as conn:
        after = {
            table: conn.execute(f"SELECT COUNT(*) AS c FROM {table}").fetchone()["c"]
            for table in before
        }
    assert after == before


def test_product_menu_export_rejects_unit_user_and_returns_valid_empty_workbook(tmp_path):
    client = make_client(tmp_path)
    admin_headers = login(client, "root_admin", "StrongPassword123")
    product = create_product(client, admin_headers, code="MENU-ONLY", name="仅测试", unit="斤", price_cents=200)
    create_unit_order(client, admin_headers, product["id"], "EMPTY", "1")
    unit_headers = login(client, "unit_EMPTY", "UnitEMPTYPassword123")

    assert client.get("/api/v1/admin/products/export.xlsx", headers=unit_headers).status_code == 403
    deleted = client.delete(f"/api/v1/admin/products/{product['id']}", headers=admin_headers)
    assert deleted.status_code == 200, deleted.text

    exported = client.get("/api/v1/admin/products/export.xlsx", headers=admin_headers)
    assert exported.status_code == 200, exported.text
    workbook = load_workbook(BytesIO(exported.content), data_only=True)
    assert workbook.sheetnames == ["三公鲜配商品菜单"]
    assert [cell.value for cell in workbook.active[1]] == ["序号", "商品名称", "规格", "当前价格"]
    assert workbook.active.max_row == 1


def test_product_menu_workbook_escapes_formula_text():
    from app.services.exports import product_menu_workbook

    workbook = load_workbook(
        BytesIO(product_menu_workbook([{"name": "=SUM(1,1)", "spec": "+危险规格", "price_cents": 250}])),
        data_only=False,
    )
    sheet = workbook.active
    assert sheet["B2"].value == "'=SUM(1,1)"
    assert sheet["C2"].value == "'+危险规格"
    assert sheet["D2"].value == 2.5
