from concurrent.futures import ThreadPoolExecutor
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from test_workflows import advance_to_preparing, login, make_client


def _create_unit_user(client, admin_headers, suffix: str):
    unit = client.post(
        "/api/v1/admin/units",
        headers=admin_headers,
        json={
            "unit_code": f"PH3-{suffix}",
            "unit_name": f"Phase3单位{suffix}",
            "default_delivery_point": f"Phase3收货点{suffix}",
        },
    )
    assert unit.status_code == 200, unit.text
    unit_id = unit.json()["id"]
    user = client.post(
        "/api/v1/admin/users",
        headers=admin_headers,
        json={
            "username": f"phase3_unit_{suffix}",
            "password": "UnitPassword123",
            "display_name": f"Phase3账号{suffix}",
            "role": "unit_user",
            "unit_id": unit_id,
            "must_change_password": False,
        },
    )
    assert user.status_code == 200, user.text
    return login(client, f"phase3_unit_{suffix}", "UnitPassword123")


def _create_product(client, admin_headers, code: str, name: str, unit: str):
    response = client.post(
        "/api/v1/admin/products",
        headers=admin_headers,
        json={
            "product_code": code,
            "name": name,
            "category": "粮油",
            "spec": "Phase3测试",
            "unit": unit,
            "price_cents": 100,
            "stock_quantity": "1000",
            "min_order_quantity": "1",
            "quantity_step": "1",
            "warning_quantity": "0",
            "supply_status": "normal",
        },
    )
    assert response.status_code == 200, response.text
    return response.json()["id"]


def _create_preparing_order(client, admin_headers, unit_headers, product_id, quantity):
    response = client.post(
        "/api/v1/orders",
        headers=unit_headers,
        json={"items": [{"product_id": product_id, "quantity": str(quantity)}]},
    )
    assert response.status_code == 200, response.text
    order = response.json()
    advance_to_preparing(client, admin_headers, order["id"])
    return order


def test_phase3_aggregates_by_product_and_unit_and_exports(tmp_path):
    client = make_client(tmp_path)
    admin_headers = login(client, "root_admin", "StrongPassword123")
    jin_product = _create_product(client, admin_headers, "PH3-POTATO-JIN", "土豆", "斤")
    box_product = _create_product(client, admin_headers, "PH3-POTATO-BOX", "土豆", "箱")

    unit_a = _create_unit_user(client, admin_headers, "A")
    unit_b = _create_unit_user(client, admin_headers, "B")
    unit_c = _create_unit_user(client, admin_headers, "C")
    orders = [
        _create_preparing_order(client, admin_headers, unit_a, jin_product, 5),
        _create_preparing_order(client, admin_headers, unit_b, jin_product, 7),
        _create_preparing_order(client, admin_headers, unit_c, jin_product, 4),
        _create_preparing_order(client, admin_headers, unit_c, box_product, 2),
    ]

    batch = client.post(
        "/api/v1/admin/batches",
        headers=admin_headers,
        json={"name": "Phase3测试备货单", "note": "隔离测试", "order_ids": [order["id"] for order in orders]},
    )
    assert batch.status_code == 200, batch.text
    batch_body = batch.json()
    assert batch_body["status"] == "open"
    assert len(batch_body["orders"]) == 4
    listed = client.get("/api/v1/admin/batches", headers=admin_headers)
    assert listed.status_code == 200, listed.text
    assert listed.json()["items"][0]["order_count"] == 4
    assert set(listed.json()["items"][0]["unit_codes"].split(",")) == {"PH3-A", "PH3-B", "PH3-C"}
    assert set(listed.json()["items"][0]["unit_labels"].split(",")) == {
        "PH3-A · Phase3单位A",
        "PH3-B · Phase3单位B",
        "PH3-C · Phase3单位C",
    }

    order_list = client.get("/api/v1/admin/orders?page=1&page_size=100", headers=admin_headers)
    assert order_list.status_code == 200, order_list.text
    assert {item["unit_code"] for item in order_list.json()["items"]} >= {"PH3-A", "PH3-B", "PH3-C"}

    summary = client.get(f"/api/v1/admin/batches/{batch_body['id']}/summary", headers=admin_headers)
    assert summary.status_code == 200, summary.text
    body = summary.json()
    jin = next(item for item in body["by_product"] if item["product_id"] == jin_product)
    box = next(item for item in body["by_product"] if item["product_id"] == box_product)
    assert jin["total_quantity"] == "16"
    assert jin["unit"] == "斤"
    assert box["total_quantity"] == "2"
    assert box["unit"] == "箱"
    assert body["unit_count"] == 3
    assert body["product_count"] == 2
    assert [unit["unit_code"] for unit in body["by_unit"]] == ["PH3-A", "PH3-B", "PH3-C"]

    export = client.get(f"/api/v1/admin/batches/{batch_body['id']}/picking-list.xlsx", headers=admin_headers)
    assert export.status_code == 200, export.text
    workbook = load_workbook(BytesIO(export.content), read_only=True)
    assert {"总计", "米面粮油"}.issubset(workbook.sheetnames)
    assert {"PH3-A-米面粮油", "PH3-B-米面粮油", "PH3-C-米面粮油"}.issubset(workbook.sheetnames)
    report_rows = [row for row in workbook["总计"].iter_rows(min_row=4, values_only=True) if isinstance(row[0], int)]
    assert {row[3] for row in report_rows} == {"斤", "箱"}

    before = client.get(f"/api/v1/products/{jin_product}", headers=unit_a).json()
    completed = client.patch(
        f"/api/v1/admin/batches/{batch_body['id']}/status",
        headers=admin_headers,
        json={"status": "closed", "expected_version": batch_body["version"]},
    )
    assert completed.status_code == 200, completed.text
    assert completed.json()["status"] == "closed"
    assert completed.json()["orders"][0]["status"] == "preparing"
    after = client.get(f"/api/v1/products/{jin_product}", headers=unit_a).json()
    assert (after["stock_quantity"], after["reserved_quantity"]) == (before["stock_quantity"], before["reserved_quantity"])


def test_phase3_rejects_pending_duplicate_and_supports_archive_reorganization(tmp_path):
    client = make_client(tmp_path)
    admin_headers = login(client, "root_admin", "StrongPassword123")
    product_id = _create_product(client, admin_headers, "PH3-REJECT", "Phase3土豆", "斤")
    unit_headers = _create_unit_user(client, admin_headers, "R")
    pending = client.post(
        "/api/v1/orders",
        headers=unit_headers,
        json={"items": [{"product_id": product_id, "quantity": "1"}]},
    ).json()
    rejected = client.post(
        "/api/v1/admin/batches",
        headers=admin_headers,
        json={"name": "不应生成", "order_ids": [pending["id"]]},
    )
    assert rejected.status_code == 409

    preparing = _create_preparing_order(client, admin_headers, unit_headers, product_id, 2)
    first = client.post(
        "/api/v1/admin/batches",
        headers=admin_headers,
        json={"name": "可归档备货单", "order_ids": [preparing["id"]]},
    )
    assert first.status_code == 200, first.text
    first_body = first.json()
    duplicate = client.post(
        "/api/v1/admin/batches",
        headers=admin_headers,
        json={"name": "重复加入", "order_ids": [preparing["id"]]},
    )
    assert duplicate.status_code == 409

    archived = client.delete(f"/api/v1/admin/batches/{first_body['id']}", headers=admin_headers)
    assert archived.status_code == 200, archived.text
    assert archived.json()["status"] == "cancelled"
    default_batches = client.get("/api/v1/admin/batches", headers=admin_headers)
    assert default_batches.status_code == 200, default_batches.text
    assert first_body["id"] not in {item["id"] for item in default_batches.json()["items"]}
    archived_batches = client.get("/api/v1/admin/batches?status=cancelled", headers=admin_headers)
    assert archived_batches.status_code == 200, archived_batches.text
    assert first_body["id"] in {item["id"] for item in archived_batches.json()["items"]}
    eligible = client.get("/api/v1/admin/batches/eligible-orders", headers=admin_headers).json()["items"]
    assert any(order["id"] == preparing["id"] for order in eligible)

    reorganized = client.post(
        "/api/v1/admin/batches",
        headers=admin_headers,
        json={"name": "重新组织备货单", "order_ids": [preparing["id"]]},
    )
    assert reorganized.status_code == 200, reorganized.text
    closed = client.patch(
        f"/api/v1/admin/batches/{reorganized.json()['id']}/status",
        headers=admin_headers,
        json={"status": "closed"},
    )
    assert closed.status_code == 200, closed.text
    cannot_delete = client.delete(f"/api/v1/admin/batches/{reorganized.json()['id']}", headers=admin_headers)
    assert cannot_delete.status_code == 409


def test_phase3_date_filter_bulk_export_and_concurrent_duplicate_protection(tmp_path):
    client = make_client(tmp_path)
    admin_headers = login(client, "root_admin", "StrongPassword123")
    product_id = _create_product(client, admin_headers, "PH3-BULK", "Phase3白菜", "斤")
    unit_headers = _create_unit_user(client, admin_headers, "BULK")
    order = _create_preparing_order(client, admin_headers, unit_headers, product_id, 3)

    client_a = client
    from app.main import app

    client_b = TestClient(app)
    headers_b = login(client_b, "root_admin", "StrongPassword123")

    def create_batch(test_client, headers):
        return test_client.post(
            "/api/v1/admin/batches",
            headers=headers,
            json={"name": "并发备货单", "order_ids": [order["id"]]},
        ).status_code

    with ThreadPoolExecutor(max_workers=2) as executor:
        statuses = sorted(executor.map(lambda args: create_batch(*args), [(client_a, admin_headers), (client_b, headers_b)]))
    assert statuses == [200, 409]

    batches = client.get("/api/v1/admin/batches", headers=admin_headers).json()["items"]
    assert len(batches) == 1
    batch_id = batches[0]["id"]
    assert client.get("/api/v1/admin/batches?date_from=2099-01-01", headers=admin_headers).json()["items"] == []
    assert client.get("/api/v1/admin/batches?date_to=2000-01-01", headers=admin_headers).json()["items"] == []

    bulk = client.get(
        f"/api/v1/admin/batches/bulk.xlsx?batch_ids={batch_id}&batch_ids={batch_id}",
        headers=admin_headers,
    )
    assert bulk.status_code == 200, bulk.text
    workbook = load_workbook(BytesIO(bulk.content), read_only=True)
    assert any(name.startswith("总计-") for name in workbook.sheetnames)
