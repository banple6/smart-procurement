from __future__ import annotations

import csv
import re
import unicodedata
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from fastapi import HTTPException
from openpyxl import load_workbook


MAX_FILE_BYTES = 10 * 1024 * 1024
MAX_SHEETS = 20
MAX_ROWS = 10_000
MAX_COLUMNS = 100
SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".csv"}

FIELD_SYNONYMS = {
    "product_code": {"编码", "商品编码", "货号", "sku", "物料编码", "产品编码"},
    "product_name": {"商品", "品名", "品类", "菜品", "货品", "货物名称", "食材", "食材名称", "商品名称", "名称"},
    "category": {"分类", "商品分类", "食材分类", "货品分类"},
    "spec": {"规格", "规格型号", "等级", "级别", "包装规格", "型号"},
    "unit": {"单位", "计量单位", "uom", "计量"},
    "stock": {"库存", "库存数量", "期初库存", "现有库存", "数量"},
    "price": {"价格", "单价", "报价", "执行价", "本期价", "本期执行价", "今日价", "结算价", "协议价", "含税单价", "采购价"},
}


def normalized_text(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).replace("\u3000", " ")
    text = "".join(char for char in text if unicodedata.category(char) != "Cf")
    text = re.sub(r"\s+", " ", text).strip().casefold()
    return text


def normalize_product_name(value: object) -> str:
    """Normalize formatting only; never infer a product synonym or specification."""
    return normalized_text(value)


def compact_label(value: object) -> str:
    """Header labels may contain visual spaces that are not part of the field name."""
    return normalized_text(value).replace(" ", "")


def display_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return unicodedata.normalize("NFKC", str(value)).strip()


def parse_price_cents(value: object) -> int | None:
    raw = display_text(value).replace(",", "")
    raw = re.sub(r"[￥¥$元人民币RMB\s]", "", raw, flags=re.IGNORECASE)
    if not raw:
        return None
    try:
        amount = Decimal(raw)
    except InvalidOperation:
        return None
    if amount <= 0 or amount > Decimal("1000000"):
        return None
    return int((amount * 100).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def price_header_unit(value: object) -> str:
    header = compact_label(value)
    match = re.search(r"(?:元|¥|￥)(?:/|／|每)?(500克|500g|公斤|千克|kg|克|g|斤|吨)", header)
    return display_text(match.group(1)) if match else ""


def unit_conversion(source_unit: str, target_unit: str, source_cents: int) -> tuple[str, int, str] | None:
    source = normalized_text(source_unit)
    target = normalized_text(target_unit)
    kilograms = {"kg", "公斤", "千克"}
    grams = {"g", "克"}
    if source == target:
        return display_text(source_unit), source_cents, "1"
    if target in kilograms:
        if source in kilograms:
            return "公斤", source_cents, "1"
        if source in grams:
            return "公斤", int((Decimal(source_cents) * Decimal("1000")).to_integral_value(rounding=ROUND_HALF_UP)), "1000"
        if source in {"500g", "500克"}:
            return "公斤", int((Decimal(source_cents) * Decimal("2")).to_integral_value(rounding=ROUND_HALF_UP)), "2"
        if source == "吨":
            return "公斤", int((Decimal(source_cents) / Decimal("1000")).to_integral_value(rounding=ROUND_HALF_UP)), "0.001"
        if source == "斤":
            return "公斤", int((Decimal(source_cents) * Decimal("2")).to_integral_value(rounding=ROUND_HALF_UP)), "2"
    if target == "斤" and source in kilograms:
        return "斤", int((Decimal(source_cents) / Decimal("2")).to_integral_value(rounding=ROUND_HALF_UP)), "0.5"
    return None


@dataclass(frozen=True)
class ParsedSheet:
    name: str
    rows: list[list[str]]


def _trim_rows(rows: list[list[str]]) -> list[list[str]]:
    while rows and not any(rows[-1]):
        rows.pop()
    if not rows:
        return []
    width = min(MAX_COLUMNS, max(len(row) for row in rows))
    result = [row[:width] + [""] * max(0, width - len(row)) for row in rows]
    while width and not any(row[width - 1] for row in result):
        width -= 1
    return [row[:width] for row in result]


def _xlsx_sheets(path: Path) -> list[ParsedSheet]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    if len(workbook.worksheets) > MAX_SHEETS:
        raise HTTPException(status_code=400, detail=f"Excel Sheet 数量不能超过 {MAX_SHEETS} 个")
    result = []
    for sheet in workbook.worksheets:
        rows = []
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_number > MAX_ROWS or len(row) > MAX_COLUMNS:
                raise HTTPException(status_code=400, detail=f"Sheet“{sheet.title}”超过 {MAX_ROWS} 行或 {MAX_COLUMNS} 列限制")
            rows.append([display_text(cell) for cell in row])
        result.append(ParsedSheet(sheet.title, _trim_rows(rows)))
    return result


def _xls_sheets(path: Path) -> list[ParsedSheet]:
    try:
        import xlrd
    except ImportError as exc:
        raise HTTPException(status_code=400, detail="服务器未安装 .xls 兼容组件，请转换为 .xlsx 后重试") from exc
    workbook = xlrd.open_workbook(path, on_demand=True)
    if workbook.nsheets > MAX_SHEETS:
        raise HTTPException(status_code=400, detail=f"Excel Sheet 数量不能超过 {MAX_SHEETS} 个")
    result = []
    for name in workbook.sheet_names():
        sheet = workbook.sheet_by_name(name)
        if sheet.nrows > MAX_ROWS or sheet.ncols > MAX_COLUMNS:
            raise HTTPException(status_code=400, detail=f"Sheet“{name}”超过 {MAX_ROWS} 行或 {MAX_COLUMNS} 列限制")
        rows = [[display_text(sheet.cell_value(r, c)) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
        result.append(ParsedSheet(name, _trim_rows(rows)))
    return result


def _csv_sheets(path: Path) -> list[ParsedSheet]:
    content = path.read_bytes()
    for encoding in ("utf-8-sig", "gb18030", "utf-8"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise HTTPException(status_code=400, detail="CSV 编码无法识别，请使用 UTF-8 或 GB18030 文件")
    rows = [[display_text(value) for value in row] for row in csv.reader(text.splitlines())]
    if len(rows) > MAX_ROWS or any(len(row) > MAX_COLUMNS for row in rows):
        raise HTTPException(status_code=400, detail=f"CSV 超过 {MAX_ROWS} 行或 {MAX_COLUMNS} 列限制")
    return [ParsedSheet("CSV", _trim_rows(rows))]


def parse_workbook(path: Path) -> list[ParsedSheet]:
    if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="仅支持 .xlsx、.xls、.csv 报价文件")
    if not path.exists() or path.stat().st_size == 0:
        raise HTTPException(status_code=400, detail="Excel 文件为空或不存在")
    if path.stat().st_size > MAX_FILE_BYTES:
        raise HTTPException(status_code=400, detail="Excel 文件不能超过 10 MB")
    try:
        if path.suffix.lower() == ".xlsx":
            sheets = _xlsx_sheets(path)
        elif path.suffix.lower() == ".xls":
            sheets = _xls_sheets(path)
        else:
            sheets = _csv_sheets(path)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Excel 文件损坏、加密或无法读取") from exc
    if not any(sheet.rows for sheet in sheets):
        raise HTTPException(status_code=400, detail="Excel 中没有可识别的数据")
    return sheets


def detect_header(rows: list[list[str]]) -> dict:
    candidates = []
    for index, row in enumerate(rows[: min(20, len(rows))]):
        mapping: dict[str, int] = {}
        price_columns: list[int] = []
        score = 0
        for col, value in enumerate(row):
            label = compact_label(value)
            if not label:
                continue
            code_column = any(compact_label(name) in label for name in FIELD_SYNONYMS["product_code"])
            for field, names in FIELD_SYNONYMS.items():
                if field == "product_name" and code_column:
                    continue
                if any(compact_label(name) in label for name in names):
                    if field == "price":
                        price_columns.append(col)
                    elif field not in mapping:
                        mapping[field] = col
                    score += 2
        if price_columns:
            preferred = [col for col in price_columns if any(word in normalized_text(row[col]) for word in ("执行", "协议", "结算", "本期", "今日"))]
            mapping["price"] = (preferred or price_columns)[0]
        if "product_name" in mapping and "price" in mapping:
            score += 8
        candidates.append({"header_row": index + 1, "mapping": mapping, "score": score, "ambiguous_price": len(price_columns) > 1})
    candidates.sort(key=lambda item: item["score"], reverse=True)
    return candidates[0] if candidates else {"header_row": None, "mapping": {}, "score": 0, "ambiguous_price": False}


def mapping_from_names(header: list[str], mapping: dict[str, str]) -> dict[str, int]:
    resolved: dict[str, int] = {}
    names = {normalized_text(value): index for index, value in enumerate(header)}
    for field, name in mapping.items():
        index = names.get(normalized_text(name))
        if index is None:
            raise HTTPException(status_code=400, detail=f"确认的字段“{name}”不在 Excel 表头中")
        resolved[field] = index
    if "product_name" not in resolved or "price" not in resolved:
        raise HTTPException(status_code=400, detail="请至少确认商品名称列和执行价格列")
    return resolved
