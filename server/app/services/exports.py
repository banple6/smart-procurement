from io import BytesIO
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from .order_status import order_status_payload


LEDGER_HEADERS = [
    "序号",
    "订单编号",
    "商品分类",
    "商品名称",
    "规格",
    "计量单位",
    "数量",
    "单价",
    "小计",
    "订单金额",
    "订单状态",
    "下单时间",
    "单位名称",
]


def ledger_workbook(rows: list[dict]) -> bytes:
    wb = Workbook()
    wb.properties.title = "三公鲜配采购台账"
    ws = wb.active
    ws.title = "订单台账"
    ws.append(LEDGER_HEADERS)
    summary: dict[tuple[str, str, str, str], dict[str, float | int | str]] = defaultdict(lambda: {"quantity": 0.0, "subtotal_cents": 0})
    for index, row in enumerate(rows, start=1):
        ws.append(
            [
                index,
                row["order_no"],
                row["category_snapshot"],
                row["product_name_snapshot"],
                row["spec_snapshot"],
                row["unit_snapshot"],
                row["quantity"],
                row["price_cents_snapshot"] / 100,
                row["subtotal_cents"] / 100,
                row["total_cents"] / 100,
                order_status_payload(row["status"])["status_label"],
                row["created_at"],
                row["unit_name_snapshot"],
            ]
        )
        key = (
            row["category_snapshot"],
            row["product_name_snapshot"],
            row["spec_snapshot"],
            row["unit_snapshot"],
        )
        summary[key]["quantity"] = float(summary[key]["quantity"]) + float(row["quantity"])
        summary[key]["subtotal_cents"] = int(summary[key]["subtotal_cents"]) + int(row["subtotal_cents"])
    summary_ws = wb.create_sheet("商品需求汇总")
    summary_ws.append(["商品分类", "商品名称", "规格", "计量单位", "需求数量", "需求金额"])
    for (category, name, spec, unit), values in sorted(summary.items(), key=lambda item: item[0][1]):
        summary_ws.append(
            [
                category,
                name,
                spec,
                unit,
                values["quantity"],
                int(values["subtotal_cents"]) / 100,
            ]
        )
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


PREPARATION_HEADERS = [
    "商品编码",
    "商品名称",
    "规格",
    "单位",
    "申领数量",
    "实际数量",
    "单位数",
    "订单数",
]


def preparation_summary_workbook(rows: list[dict]) -> bytes:
    wb = Workbook()
    wb.properties.title = "三公鲜配今日备货单"
    ws = wb.active
    ws.title = "今日备货汇总"
    ws.append(PREPARATION_HEADERS)
    for row in rows:
        ws.append(
            [
                row["product_code"],
                row["product_name"],
                row["spec"],
                row["unit"],
                row["requested_quantity"],
                row["actual_quantity"],
                row["unit_count"],
                row["order_count"],
            ]
        )
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def delivery_sheets_workbook(units: list[dict]) -> bytes:
    wb = Workbook()
    wb.properties.title = "三公鲜配配送单"
    summary = wb.active
    summary.title = "按商品汇总"
    summary.append(PREPARATION_HEADERS)
    product_totals: dict[tuple[str, str, str, str], dict] = defaultdict(
        lambda: {"requested_quantity": 0.0, "actual_quantity": 0.0, "unit_ids": set(), "order_ids": set()}
    )
    for unit in units:
        for order in unit["orders"]:
            for item in order["items"]:
                key = (item["product_code"], item["product_name"], item["spec"], item["unit"])
                product_totals[key]["requested_quantity"] += float(item["requested_quantity"])
                product_totals[key]["actual_quantity"] += float(item["actual_quantity"])
                product_totals[key]["unit_ids"].add(unit["unit_id"])
                product_totals[key]["order_ids"].add(order["order_id"])
    for (code, name, spec, unit_name), values in sorted(product_totals.items(), key=lambda item: item[0][1]):
        summary.append(
            [
                code,
                name,
                spec,
                unit_name,
                values["requested_quantity"],
                values["actual_quantity"],
                len(values["unit_ids"]),
                len(values["order_ids"]),
            ]
        )

    detail = wb.create_sheet("按单位配送")
    detail.append(["单位名称", "配送点", "订单编号", "状态", "商品编码", "商品名称", "规格", "单位", "申领数量", "实际数量", "调整原因"])
    for unit in units:
        for order in unit["orders"]:
            for item in order["items"]:
                detail.append(
                    [
                        unit["unit_name"],
                        unit["delivery_point"],
                        order["order_no"],
                        order["status"],
                        item["product_code"],
                        item["product_name"],
                        item["spec"],
                        item["unit"],
                        item["requested_quantity"],
                        item["actual_quantity"],
                        item["adjustment_reason"],
                    ]
                )
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def product_import_template_workbook() -> bytes:
    wb = Workbook()
    wb.properties.title = "三公鲜配食材导入模板"
    ws = wb.active
    ws.title = "食材导入"
    headers = ["商品编码", "商品名称", "商品分类", "规格", "计量单位", "单价（元）", "库存"]
    ws.append(headers)
    ws.append(["", "示例：大白菜", "蔬菜", "散装", "斤", 0.56, 0])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F5A94")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    widths = [18, 24, 16, 18, 14, 16, 14]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index)].width = width
    ws.column_dimensions["F"].number_format = "0.00"
    category_validation = DataValidation(
        type="list",
        formula1='"蔬菜,水果,肉禽,水产,粮油,蛋奶,调料,其他"',
        allow_blank=True,
    )
    unit_validation = DataValidation(
        type="list",
        formula1='"公斤,斤,箱,袋,个,筐,盒,瓶,份,包"',
        allow_blank=False,
    )
    ws.add_data_validation(category_validation)
    ws.add_data_validation(unit_validation)
    category_validation.add("C2:C1000")
    unit_validation.add("E2:E1000")

    guide = wb.create_sheet("填写说明")
    guide.append(["字段", "说明"])
    guide.append(["商品名称", "必填，用于与系统现有食材按标准化名称精确匹配"])
    guide.append(["商品编码", "选填；填写后优先按编码精确匹配"])
    guide.append(["商品分类", "选填；缺失时可在导入审核页统一设置"])
    guide.append(["规格", "选填；公斤/斤默认散装，其他单位默认预包装"])
    guide.append(["计量单位", "必填；箱、袋等无法确定重量时不会自动换算"])
    guide.append(["单价（元）", "必填，使用十进制元金额，例如 3.50"])
    guide.append(["库存", "选填；未填写时新商品库存为 0"])
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 72
    for cell in guide[1]:
        cell.font = Font(bold=True)
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()
