from collections import OrderedDict
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .local_time import display_local_time


HEADER_FILL = PatternFill("solid", fgColor="DDEAF6")
TITLE_FILL = PatternFill("solid", fgColor="1F5A94")
THIN_BORDER = Border(
    left=Side(style="thin", color="B8C5D1"),
    right=Side(style="thin", color="B8C5D1"),
    top=Side(style="thin", color="B8C5D1"),
    bottom=Side(style="thin", color="B8C5D1"),
)


def _safe_sheet_title(value: str) -> str:
    title = "".join("_" if char in "[]:*?/\\" else char for char in (value or "其他"))
    return title[:31] or "其他"


def _unique_sheet_title(wb, value: str) -> str:
    base = _safe_sheet_title(value)
    candidate = base
    suffix = 2
    while candidate in wb.sheetnames:
        marker = f"-{suffix}"
        candidate = f"{base[:31 - len(marker)]}{marker}"
        suffix += 1
    return candidate


def document_date(value: str | None) -> str:
    localized = display_local_time(value)
    return localized[:10].replace("-", "") if len(localized) >= 10 else "日期未记录"


def business_document_date(document: dict) -> str:
    value = str(document.get("business_date") or "")
    if len(value) >= 10:
        return value[:10].replace("-", "")
    return document_date(document.get("created_at"))


def _safe_filename_part(value: str) -> str:
    return "".join("_" if char in '/\\:*?\"<>|' else char for char in str(value or "").strip()) or "未命名"



# ── Report category mapping ──────────────────────────────────────
# Maps database categories to the customer-facing report categories.
# Categories not listed here are collected under "其他".
REPORT_CATEGORY_MAP = {
    "蔬菜": "蔬菜",
    "水果": "水果",
    "肉禽": "肉",
    "肉": "肉",
    "粮油": "米面粮油",
    "米面粮油": "米面粮油",
    "调料": "调料",
    "蛋奶": "蛋奶",
    "水产": "水产",
}

# Fixed sheet order for customer-facing reports.
REPORT_CATEGORY_ORDER = ["蔬菜", "水果", "肉", "米面粮油", "调料", "蛋奶", "水产"]


def _report_category(db_category: str) -> str:
    """Map a database category to the customer report category."""
    return REPORT_CATEGORY_MAP.get(db_category or "", "其他")


def _report_category_lines(lines: list[dict]) -> OrderedDict[str, list[dict]]:
    """Group lines by report category in fixed sheet order."""
    buckets: dict[str, list[dict]] = {}
    for line in sorted(lines, key=lambda item: (item.get("category") or "其他", item.get("product_name") or "", item.get("unit") or "", item.get("price_cents") or 0)):
        report_cat = _report_category(line.get("category"))
        buckets.setdefault(report_cat, []).append(line)
    result: OrderedDict[str, list[dict]] = OrderedDict()
    for cat in REPORT_CATEGORY_ORDER:
        if cat in buckets:
            result[cat] = buckets.pop(cat)
    # Merge remaining into "其他"
    other: list[dict] = []
    for remaining_lines in buckets.values():
        other.extend(remaining_lines)
    if other:
        other.sort(key=lambda item: (item.get("category") or "其他", item.get("product_name") or ""))
        result["其他"] = other
    return result


def _setup_sheet(ws, title: str, headers: list[str], widths: list[int]):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(1, 1, title)
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    for index, header in enumerate(headers, start=1):
        cell = ws.cell(3, index, header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[chr(64 + index)].width = widths[index - 1]
    ws.freeze_panes = "A4"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.print_title_rows = "1:3"


def _style_data_rows(ws, start_row: int, end_row: int, column_count: int):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=column_count):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def _category_lines(lines: list[dict]) -> OrderedDict[str, list[dict]]:
    categories: OrderedDict[str, list[dict]] = OrderedDict()
    for line in sorted(lines, key=lambda item: (item.get("category") or "其他", item.get("product_name") or "", item.get("unit") or "", item.get("price_cents") or 0)):
        categories.setdefault(line.get("category") or "其他", []).append(line)
    return categories


def _picking_title_unit(unit: dict) -> str:
    return str(unit.get("unit_name") or "").strip() or str(unit.get("unit_code") or "未编码")


def _append_signatures(ws, row: int, outbound: bool = False):
    row += 2
    if outbound:
        ws.cell(row, 1, "发货人签字：________________")
        ws.cell(row, 3, "接收人签字：________________")
        ws.cell(row, 6, "日期：________________")
    else:
        ws.cell(row, 1, "接收人签字：________________")
        ws.cell(row, 4, "日期：________________")
    ws.row_dimensions[row].height = 28


def _picking_sheet(
    ws,
    title: str,
    lines: list[dict],
    *,
    batch_no: str = "",
):
    headers = ["序号", "商品名称", "规格", "计量单位", "需求数量"]
    _setup_sheet(ws, title, headers, [10, 24, 20, 14, 16])
    ws.merge_cells("A2:C2")
    ws.merge_cells("D2:E2")
    ws.cell(2, 4, f"系统备货单号：{batch_no}")
    ws.cell(2, 4).alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
    ws.row_dimensions[2].height = 22
    for index, line in enumerate(lines, start=1):
        ws.append([index, line["product_name"], line.get("spec") or "", line["unit"], line["actual_quantity"]])
    end_row = 3 + len(lines)
    if lines:
        _style_data_rows(ws, 4, end_row, len(headers))
        for cell in ws["E"][3:end_row]:
            cell.number_format = "0.###"
    _append_signatures(ws, end_row)
    ws.print_area = f"A1:E{end_row + 2}"


def _append_unit_picking_sheets(wb, aggregation: dict, name_prefix: str = "", use_category_sheet_names: bool = False):
    batch = aggregation["batch"]
    date = business_document_date(batch)
    for unit in aggregation["by_unit"]:
        code = unit.get("unit_code") or "未编码"
        title_unit = _picking_title_unit(unit)
        for category, lines in _report_category_lines(unit["items"]).items():
            sheet_name = f"{name_prefix}{category}" if use_category_sheet_names else f"{name_prefix}{code}-{category}"
            ws = wb.create_sheet(_unique_sheet_title(wb, sheet_name))
            _picking_sheet(
                ws,
                f"{category}备货单（{date}/{title_unit}）",
                lines,
                batch_no=batch["batch_no"],
            )


def _picking_summary_title(aggregation: dict) -> str:
    batch = aggregation["batch"]
    if len(aggregation["by_unit"]) == 1:
        return f"三公鲜配备货单（{business_document_date(batch)}/{_picking_title_unit(aggregation['by_unit'][0])}）"
    return f"三公鲜配备货单（{batch['batch_no']}）"


def batch_picking_workbook(aggregation: dict) -> bytes:
    lines = aggregation["document_lines"]
    categories = _report_category_lines(lines)
    wb = Workbook()
    wb.properties.title = f"三公鲜配备货单 {aggregation['batch']['batch_no']}"
    wb.remove(wb.active)
    _append_unit_picking_sheets(wb, aggregation, use_category_sheet_names=len(aggregation["by_unit"]) == 1)
    total = wb.create_sheet("总计")
    _picking_sheet(total, _picking_summary_title(aggregation), lines, batch_no=aggregation["batch"]["batch_no"])
    if len(aggregation["by_unit"]) > 1:
        for category, category_lines in categories.items():
            ws = wb.create_sheet(_unique_sheet_title(wb, category))
            _picking_sheet(ws, f"{category}备货单（{aggregation['batch']['batch_no']}）", category_lines, batch_no=aggregation["batch"]["batch_no"])
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def batch_picking_workbook_multi(aggregations: list[dict]) -> bytes:
    """Export selected preparation orders into one workbook for easy saving."""
    wb = Workbook()
    wb.remove(wb.active)
    for aggregation in aggregations:
        batch_no = aggregation["batch"]["batch_no"]
        short_batch = batch_no[-4:]
        _append_unit_picking_sheets(wb, aggregation, name_prefix=f"{short_batch}-")
        total = wb.create_sheet(_unique_sheet_title(wb, f"总计-{batch_no}"))
        _picking_sheet(total, _picking_summary_title(aggregation), aggregation["document_lines"], batch_no=batch_no)
        if len(aggregation["by_unit"]) > 1:
            for category, category_lines in _report_category_lines(aggregation["document_lines"]).items():
                ws = wb.create_sheet(_unique_sheet_title(wb, f"{batch_no}-{category}"))
                _picking_sheet(ws, f"{category}备货单（{batch_no}）", category_lines, batch_no=batch_no)
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _outbound_sheet(ws, title: str, lines: list[dict], include_category: bool):
    headers = (["序号", "商品分类"] if include_category else ["序号"]) + ["商品名称", "计量单位", "需求数量", "单价（元）", "小计（元）"]
    widths = ([10, 16] if include_category else [10]) + [24, 14, 16, 16, 18]
    _setup_sheet(ws, title, headers, widths)
    total_cents = 0
    for index, line in enumerate(lines, start=1):
        values = [index]
        if include_category:
            values.append(line.get("category") or "其他")
        values.extend(
            [
                line["product_name"],
                line["unit"],
                line["actual_quantity"],
                Decimal(int(line["price_cents"])) / Decimal(100),
                Decimal(int(line["subtotal_cents"])) / Decimal(100),
            ]
        )
        ws.append(values)
        total_cents += int(line["subtotal_cents"])
    end_row = 3 + len(lines)
    if lines:
        _style_data_rows(ws, 4, end_row, len(headers))
        for row in range(4, end_row + 1):
            ws.cell(row, len(headers) - 2).number_format = "0.###"
            ws.cell(row, len(headers) - 1).number_format = '¥0.00'
            ws.cell(row, len(headers)).number_format = '¥0.00'
    total_row = end_row + 1
    ws.cell(total_row, len(headers) - 1, "合计")
    ws.cell(total_row, len(headers), Decimal(total_cents) / Decimal(100))
    ws.cell(total_row, len(headers)).number_format = '¥0.00'
    ws.cell(total_row, len(headers) - 1).font = Font(bold=True)
    ws.cell(total_row, len(headers)).font = Font(bold=True)
    _append_signatures(ws, total_row, outbound=True)


def batch_outbound_workbook(aggregation: dict) -> bytes:
    lines = aggregation["document_lines"]
    categories = _report_category_lines(lines)
    wb = Workbook()
    wb.properties.title = f"三公鲜配出库单 {aggregation['batch']['batch_no']}"
    total = wb.active
    total.title = "总计"
    _outbound_sheet(total, f"三公鲜配出库单（{aggregation['batch']['batch_no']}）", lines, include_category=True)
    category_summary_start = 7 + len(lines)
    total.cell(category_summary_start, 1, "分类金额汇总")
    total.cell(category_summary_start, 1).font = Font(bold=True)
    for offset, (category, category_lines) in enumerate(categories.items(), start=1):
        total.cell(category_summary_start + offset, 1, category)
        category_cents = sum(int(line["subtotal_cents"]) for line in category_lines)
        total.cell(category_summary_start + offset, 2, Decimal(category_cents) / Decimal(100))
        total.cell(category_summary_start + offset, 2).number_format = '¥0.00'
    for category, category_lines in categories.items():
        ws = wb.create_sheet(_safe_sheet_title(category))
        _outbound_sheet(ws, f"{category}出库单（{aggregation['batch']['batch_no']}）", category_lines, include_category=False)
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def outbound_order_workbook(outbound: dict, lines: list[dict]) -> bytes:
    """Create one printable outbound document for one receiving unit."""
    wb = Workbook()
    wb.properties.title = f"三公鲜配出库单 {outbound['outbound_no']}"
    ws = wb.active
    ws.title = "出库单"
    headers = ["序号", "食品分类", "食材名称", "规格", "计量单位", "需求数量"]
    code = outbound.get("unit_code") or "未编码"
    unit_name = str(outbound.get("unit_name_snapshot") or "").strip()
    title_unit = unit_name or code
    date = business_document_date(outbound)
    _setup_sheet(ws, f"三公鲜配出库单（{date}/{title_unit}）", headers, [8, 16, 24, 20, 14, 16])

    # Keep the document metadata readable without relying on a manual column resize.
    ws.merge_cells("A2:B2")
    ws.merge_cells("C2:D2")
    ws.merge_cells("E2:F2")
    ws.cell(2, 1, f"单位：{code} · {unit_name or '未命名单位'}")
    ws.cell(2, 3, f"系统出库单号：{outbound['outbound_no']}")
    ws.cell(2, 5, f"日期：{outbound.get('created_at') or ''}")
    for column in (1, 3, 5):
        ws.cell(2, column).alignment = Alignment(vertical="center", wrap_text=False)
    ws.row_dimensions[2].height = 22

    total_cents = sum(int(line.get("subtotal_cents") or 0) for line in lines)
    for index, line in enumerate(lines, start=1):
        ws.append([
            index,
            _report_category(line.get("category") or ""),
            line["product_name"],
            line.get("spec") or "",
            line["unit"],
            line["quantity"],
        ])
    end_row = 3 + len(lines)
    if lines:
        _style_data_rows(ws, 4, end_row, len(headers))
        for cell in ws["F"][3:end_row]:
            cell.number_format = "0.###"

    total_row = end_row + 1
    total_label = ws.cell(total_row, 5, "总金额：")
    total_value = ws.cell(total_row, 6, Decimal(total_cents) / Decimal(100))
    total_border = Border(top=Side(style="thin", color="6D7B88"))
    for cell in (total_label, total_value):
        cell.font = Font(bold=True)
        cell.border = total_border
    total_label.alignment = Alignment(horizontal="right", vertical="center")
    total_value.alignment = Alignment(horizontal="right", vertical="center")
    total_value.number_format = '¥0.00'

    signature_row = total_row + 2
    ws.merge_cells(start_row=signature_row, start_column=1, end_row=signature_row, end_column=3)
    ws.merge_cells(start_row=signature_row, start_column=4, end_row=signature_row, end_column=6)
    ws.cell(signature_row, 1, "配送人：________________")
    ws.cell(signature_row, 4, "收货人：________________")
    ws.row_dimensions[signature_row].height = 28
    ws.print_area = f"A1:F{signature_row}"
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def batch_picking_filename(aggregation: dict) -> str:
    units = aggregation.get("by_unit") or []
    if len(units) == 1 and units[0].get("unit_code"):
        unit = units[0]
        categories = list(_report_category_lines(unit.get("items") or []))
        document_name = f"{categories[0]}备货单" if len(categories) == 1 else "备货单"
        return (
            f"{document_name}_{_safe_filename_part(unit['unit_code'])}_"
            f"{_safe_filename_part(unit['unit_name'])}_{business_document_date(aggregation['batch'])}.xlsx"
        )
    return f"三公鲜配_备货单_{_safe_filename_part(aggregation['batch']['batch_no'])}.xlsx"


def outbound_order_filename(outbound: dict) -> str:
    code = outbound.get("unit_code") or "未编码"
    return (
        f"出库单_{_safe_filename_part(code)}_{_safe_filename_part(outbound['unit_name_snapshot'])}_"
        f"{business_document_date(outbound)}.xlsx"
    )


def batch_summary_workbook(aggregation: dict) -> bytes:
    wb = Workbook()
    wb.properties.title = f"三公鲜配批次汇总 {aggregation['batch']['batch_no']}"
    units = aggregation["by_unit"]
    ws = wb.active
    ws.title = "按食材汇总"
    headers = ["商品分类", "商品名称", "规格", "计量单位", "总需求"] + [unit["unit_name"] for unit in units]
    _setup_sheet(ws, f"批次食材汇总（{aggregation['batch']['batch_no']}）", headers, [16, 24, 20, 14, 16] + [16] * len(units))
    for product in aggregation["by_product"]:
        breakdown = {item["unit_id"]: item["actual_quantity"] for item in product["unit_breakdown"]}
        ws.append(
            [product["category"], product["product_name"], product["spec"], product["unit"], product["actual_quantity"]]
            + [breakdown.get(unit["unit_id"], "") for unit in units]
        )
    if aggregation["by_product"]:
        _style_data_rows(ws, 4, 3 + len(aggregation["by_product"]), len(headers))

    unit_ws = wb.create_sheet("按单位汇总")
    unit_headers = ["单位名称", "配送点", "商品分类", "商品名称", "规格", "计量单位", "需求数量"]
    _setup_sheet(unit_ws, f"批次单位汇总（{aggregation['batch']['batch_no']}）", unit_headers, [22, 26, 16, 24, 20, 14, 16])
    row_count = 0
    for unit in units:
        for item in unit["items"]:
            unit_ws.append([unit["unit_name"], unit["delivery_point"], item["category"], item["product_name"], item["spec"], item["unit"], item["actual_quantity"]])
            row_count += 1
    if row_count:
        _style_data_rows(unit_ws, 4, 3 + row_count, len(unit_headers))
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()
